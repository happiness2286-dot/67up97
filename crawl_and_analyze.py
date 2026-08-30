import urllib.request
import urllib.parse
import re
import json
import datetime
import sys
import os
from collections import Counter, defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.stdout.reconfigure(encoding='utf-8')

dow_names = {
    0: 'Thứ Hai',
    1: 'Thứ Ba',
    2: 'Thứ Tư',
    3: 'Thứ Năm',
    4: 'Thứ Sáu',
    5: 'Thứ Bảy',
    6: 'Chủ Nhật'
}

def parse_date(date_str):
    parts = date_str.split('ngày ')
    if len(parts) > 1:
        d_str = parts[1].strip()
        return datetime.datetime.strptime(d_str, "%d-%m-%Y")
    m = re.search(r'(\d{2})-(\d{2})-(\d{4})', date_str)
    if m:
        return datetime.datetime.strptime(f"{m.group(3)}-{m.group(2)}-{m.group(1)}", "%Y-%m-%d")
    m2 = re.search(r'(\d{4})-(\d{2})-(\d{2})', date_str)
    if m2:
        return datetime.datetime.strptime(date_str, "%Y-%m-%d")
    return None

def get_tong(num_str):
    if not num_str or len(num_str) < 2:
        return None
    return (int(num_str[-2]) + int(num_str[-1])) % 10

def get_bong(tong):
    if tong is None: return 0
    return (tong + 5) % 10

def get_set_20(g7_val):
    t = get_tong(g7_val)
    if t is None:
        return set()
    b = get_bong(t)
    return set(f"{i:02d}" for i in range(100) if (i//10 + i%10)%10 in (t, b))

def get_cham_g7(g7_1, g7_2, g7_3, g7_4):
    chams = set()
    for g in [g7_1, g7_2, g7_3, g7_4]:
        for char in g:
            if char.isdigit():
                chams.add(int(char))
    return chams

def get_g7_cham(g7_1, g7_2, g7_3, g7_4):
    digits = set()
    for g in [g7_1, g7_2, g7_3, g7_4]:
        for ch in g:
            if ch.isdigit():
                d = int(ch)
                digits.add(d)
                digits.add(get_bong(d))
    return digits

def filter_dan_by_head_trend(dan_set, last_de):
    """
    Quy tắc đảo chiều Đầu Lớn (5-9) vs Đầu Nhỏ (0-4):
    Nếu Đề hôm trước ra Đầu Lớn (>= 5) -> Lọc giữ lại các số Đầu Nhỏ (< 5) trong dàn.
    Nếu Đề hôm trước ra Đầu Nhỏ (< 5) -> Lọc giữ lại các số Đầu Lớn (>= 5) trong dàn.
    """
    if not last_de or len(last_de) < 2 or not last_de[0].isdigit():
        return set(dan_set)
    
    last_head = int(last_de[0])
    filtered = set()
    for num in dan_set:
        if len(num) >= 2 and num[0].isdigit():
            h = int(num[0])
            if last_head >= 5 and h < 5:
                filtered.add(num)
            elif last_head < 5 and h >= 5:
                filtered.add(num)
    return filtered if len(filtered) >= 10 else set(dan_set)

def extract_song_thu_lo_roi(dan_set, c_row):
    """
    Đánh dấu 2 con số trong dàn gốc có 'nhịp' xuất hiện/trùng lặp với bảng KQXS ngày hôm trước (Lô rơi/Đề rơi).
    """
    lottery_numbers = set()
    for k in ['db', 'g7_1', 'g7_2', 'g7_3', 'g7_4']:
        val = str(c_row.get(k, ''))
        if len(val) >= 2:
            lottery_numbers.add(val[-2:])
    
    candidates = [num for num in sorted(list(dan_set)) if num in lottery_numbers]
    if len(candidates) < 2:
        for num in sorted(list(dan_set)):
            if num not in candidates:
                candidates.append(num)
            if len(candidates) >= 2:
                break
    return candidates[:2]

def get_gaussian_rhythm_score(nhip):
    if 3 <= nhip <= 7: return 15.0
    elif 1 <= nhip <= 2: return 12.0
    elif nhip == 0: return 10.0
    elif 8 <= nhip <= 20: return 6.0
    else: return 1.0

def get_super_filtered_dan(dan_set, n1_head_freq, n1_tail_freq, n1_sum_freq, consensus_score_map=None, target_size=36):
    """
    Tạo Dàn Siêu Lọc (chỉ 25-30 số) cho N2 và N3:
    Dựa trên ma trận giao thoa từ các khung N1 trúng (Tần suất Đầu/Đuôi/Tổng xuất hiện ở N1).
    """
    if len(dan_set) <= target_size:
        return set(dan_set)
    
    consensus_score_map = consensus_score_map or {}
    scored_items = []
    
    for num in dan_set:
        if len(num) < 2 or not num.isdigit():
            continue
        h_digit = num[0]
        t_digit = num[1]
        sum_val = (int(h_digit) + int(t_digit)) % 10
        
        h_score = n1_head_freq.get(h_digit, 0) * 2.0
        t_score = n1_tail_freq.get(t_digit, 0) * 2.0
        s_score = n1_sum_freq.get(sum_val, 0) * 1.5
        c_score = consensus_score_map.get(num, 0.0)
        
        total_score = h_score + t_score + s_score + c_score
        scored_items.append((num, total_score))
        
    scored_items.sort(key=lambda x: x[1], reverse=True)
    super_set = set(x[0] for x in scored_items[:target_size])
    return super_set

def crawl_xsmb():
    url = "https://mketqua.net/so-ket-qua"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Content-Type': 'application/x-www-form-urlencoded'
    }

    data = urllib.parse.urlencode({'code': 'mb', 'count': '300', 'dow': '7'}).encode('utf-8')
    req = urllib.request.Request(url, data=data, headers=headers)

    print("Fetching lottery data from mketqua.net...", flush=True)
    try:
        with urllib.request.urlopen(req, timeout=10) as response:
            html = response.read().decode('utf-8')
            blocks = html.split('<table class="table table-condensed kqcenter kqvertimarginw table-kq-border table-kq-hover-div table-bordered kqbackground table-kq-bold-border tb-phoi-border watermark table-striped" id="result_tab_mb">')
            
            results = []
            for block in blocks[1:]:
                date_match = re.search(r'id="result_date">([^<]+)</span>', block)
                date_str = date_match.group(1).strip() if date_match else ""
                
                db_match = re.search(r'id="rs_0_0"[^>]*>(\d{5})</div>', block)
                if not db_match:
                    db_match = re.search(r'id="rs_0_0"[^>]*data-sofar="(\d{5})"', block)
                db = db_match.group(1).strip() if db_match else ""
                
                g7_1_match = re.search(r'id="rs_7_0"[^>]*>(\d{2})</div>', block)
                g7_2_match = re.search(r'id="rs_7_1"[^>]*>(\d{2})</div>', block)
                g7_3_match = re.search(r'id="rs_7_2"[^>]*>(\d{2})</div>', block)
                g7_4_match = re.search(r'id="rs_7_3"[^>]*>(\d{2})</div>', block)
                
                g7_1 = g7_1_match.group(1).strip() if g7_1_match else ""
                g7_2 = g7_2_match.group(1).strip() if g7_2_match else ""
                g7_3 = g7_3_match.group(1).strip() if g7_3_match else ""
                g7_4 = g7_4_match.group(1).strip() if g7_4_match else ""
                
                raw_prizes = re.findall(r'id="rs_\d+_\d+"[^>]*>(\d+)</div>', block)
                if not raw_prizes:
                    raw_prizes = re.findall(r'id="rs_\d+_\d+"[^>]*data-sofar="(\d+)"', block)
                all_lo = [p[-2:] for p in raw_prizes if len(p) >= 2]

                if date_str and db:
                    results.append({
                        'date': date_str,
                        'db': db,
                        'de': db[-2:] if len(db)>=2 else "",
                        'g7_1': g7_1,
                        'g7_2': g7_2,
                        'g7_3': g7_3,
                        'g7_4': g7_4,
                        'all_lo': all_lo
                    })
                    
            results_2026 = [r for r in results if '2026' in r['date']]
            print(f"Extracted {len(results_2026)} records for 2026.")
            return results_2026

    except Exception as e:
        print("Error during crawl:", e)
        return []

def analyze_all(data_2026):
    chrono = list(reversed(data_2026))
    
    for r in chrono:
        dt = parse_date(r['date'])
        if dt:
            r['datetime'] = dt
            r['iso_year'], r['iso_week'], r['iso_weekday'] = dt.isocalendar()
            r['dow_code'] = dt.weekday()
            r['dow_name'] = dow_names[dt.weekday()]

    # 1. Weekly grouping
    weekly_groups = defaultdict(list)
    for r in chrono:
        if 'iso_week' in r:
            weekly_groups[(r['iso_year'], r['iso_week'])].append(r)

    sorted_weeks = sorted(weekly_groups.keys())
    last_week_key = sorted_weeks[-2] if len(sorted_weeks) >= 2 else sorted_weeks[-1]
    last_week_records = weekly_groups[last_week_key]
    
    w_hits = {'g7_1': 0, 'g7_2': 0, 'g7_3': 0, 'g7_4': 0}
    for idx_w in range(len(last_week_records) - 1):
        target_de = last_week_records[idx_w + 1]['de']
        for k in ['g7_1', 'g7_2', 'g7_3', 'g7_4']:
            if target_de in get_set_20(last_week_records[idx_w][k]):
                w_hits[k] += 1
                
    sorted_weekly_g7 = sorted(w_hits.items(), key=lambda x: x[1], reverse=True)
    top_weekly_keys = [x[0] for x in sorted_weekly_g7[:4]]
    excluded_weekly_key = None

    # 2. Multi-window frequencies (30, 60, 90 days)
    last_date = chrono[-1]['datetime'] if 'datetime' in chrono[-1] else parse_date(chrono[-1]['date'])
    
    window_stats = {}
    for w_days in [30, 60, 90]:
        cutoff = last_date - datetime.timedelta(days=w_days)
        sub = [r for r in chrono if r.get('datetime', datetime.datetime.min) >= cutoff]
        sub_des = [r['de'] for r in sub]
        freq = Counter(sub_des)
        
        g7_sub_hits = {'g7_1': 0, 'g7_2': 0, 'g7_3': 0, 'g7_4': 0}
        for idx_s in range(len(sub) - 1):
            t_de = sub[idx_s + 1]['de']
            for k in ['g7_1', 'g7_2', 'g7_3', 'g7_4']:
                if t_de in get_set_20(sub[idx_s][k]):
                    g7_sub_hits[k] += 1
                    
        sorted_sub_g7 = sorted(g7_sub_hits.items(), key=lambda x: x[1], reverse=True)
        window_stats[w_days] = {
            'freq': freq,
            'g7_ranking': [
                {'position': k.upper(), 'hits': v, 'rate': round(v/(len(sub)-1)*100, 1) if len(sub)>1 else 0}
                for k, v in sorted_sub_g7
            ]
        }

    # 3. DAY-OF-WEEK ANALYSIS & 4 ADVANCED ALGORITHMS
    dow_stats = {}
    total_evals_2026 = len(chrono) - 1
    overall_g7_stats = {g: {'td': 0, 'b': 0} for g in ['g7_1', 'g7_2', 'g7_3', 'g7_4']}
    overall_adv_stats = {
        'cham_hits': 0,
        'goc_hits': 0,
        'inter_hits': 0,
        'stl_goc_hits': 0,
        'stl_goc_total_nhay': 0,
        'goc_lo_total_nhay': 0,
        'inter_lo_total_nhay': 0,
        'total_evals': total_evals_2026
    }
    adv_daily_records = []

    for idx in range(len(chrono) - 1):
        cur = chrono[idx]
        next_row = chrono[idx + 1]
        dt = cur.get('datetime') or parse_date(cur['date'])
        dow_code = dt.weekday() if dt else 0
        
        next_de = next_row['de']
        next_lo_list = next_row.get('all_lo', [next_de] if next_de else [])
        next_lo_counts = Counter(next_lo_list)
        
        next_de_int = int(next_de) if next_de.isdigit() else -1
        next_de_tens = next_de_int // 10
        next_de_units = next_de_int % 10
        next_de_tong = get_tong(next_row['de'])
        
        # Chams (Algo 2)
        chams = get_cham_g7(cur['g7_1'], cur['g7_2'], cur['g7_3'], cur['g7_4'])
        cham_hit = (next_de_tens in chams) or (next_de_units in chams)
        
        # Corner Pair (Algo 3)
        head_g7_1 = int(cur['g7_1'][0]) if cur['g7_1'] and cur['g7_1'][0].isdigit() else 0
        tail_g7_4 = int(cur['g7_4'][1]) if cur['g7_4'] and len(cur['g7_4'])>1 and cur['g7_4'][1].isdigit() else 0
        t_goc = (head_g7_1 + tail_g7_4) % 10
        b_goc = get_bong(t_goc)
        goc_hit = (next_de_tong in (t_goc, b_goc))
        
        # STL Cầu Ghép Góc (Lô)
        stl_p1 = f"{head_g7_1}{tail_g7_4}"
        stl_p2 = f"{tail_g7_4}{head_g7_1}"
        stl_goc_pair_str = f"{stl_p1} - {stl_p2}" if stl_p1 != stl_p2 else stl_p1
        stl_nhay = next_lo_counts[stl_p1] + (next_lo_counts[stl_p2] if stl_p2 != stl_p1 else 0)
        stl_goc_hit = stl_nhay > 0
        
        # Intersected Set (Algo 4)
        all_sums = set()
        for g in ['g7_1', 'g7_2', 'g7_3', 'g7_4']:
            t = get_tong(cur[g])
            if t is not None:
                all_sums.add(t)
                all_sums.add(get_bong(t))
            
        inter_set = set()
        for num in range(100):
            d_tens = num // 10
            d_units = num % 10
            d_sum = (d_tens + d_units) % 10
            if (d_tens in chams or d_units in chams) and (d_sum in all_sums):
                inter_set.add(f"{num:02d}")
                
        inter_hit = (next_de in inter_set)
        inter_lo_nhay = sum(next_lo_counts[n] for n in inter_set)
        
        set_20_goc = set()
        for num in range(100):
            d_sum = (num // 10 + num % 10) % 10
            if d_sum in (t_goc, b_goc):
                set_20_goc.add(f"{num:02d}")
        goc_lo_nhay = sum(next_lo_counts[n] for n in set_20_goc)
        
        if cham_hit: overall_adv_stats['cham_hits'] += 1
        if goc_hit: overall_adv_stats['goc_hits'] += 1
        if inter_hit: overall_adv_stats['inter_hits'] += 1
        if stl_goc_hit: overall_adv_stats['stl_goc_hits'] += 1
        overall_adv_stats['stl_goc_total_nhay'] += stl_nhay
        overall_adv_stats['goc_lo_total_nhay'] += goc_lo_nhay
        overall_adv_stats['inter_lo_total_nhay'] += inter_lo_nhay
        
        adv_daily_records.append({
            'stt': idx + 1,
            'date': cur['date'],
            'dow_name': dow_names[dow_code],
            'g7_1': cur['g7_1'],
            'g7_2': cur['g7_2'],
            'g7_3': cur['g7_3'],
            'g7_4': cur['g7_4'],
            'chams': "".join(str(c) for c in sorted(list(chams))),
            'sums': ", ".join(f"T{s}" for s in sorted(list(all_sums))),
            'corner_sum': f"T{t_goc}-B{b_goc}",
            'inter_size': f"{len(inter_set)} con",
            'next_date': next_row['date'],
            'next_de': next_de,
            'hit_cham': "TRÚNG" if cham_hit else "KHÔNG",
            'hit_goc': "TRÚNG" if goc_hit else "KHÔNG",
            'hit_inter': "TRÚNG" if inter_hit else "KHÔNG",
            'stl_goc_pair': stl_goc_pair_str,
            'hit_stl_goc': f"{stl_nhay} nháy 🎯" if stl_nhay > 0 else "0 nháy",
            'hit_goc_lo': f"{goc_lo_nhay} nháy",
            'hit_inter_lo': f"{inter_lo_nhay} nháy"
        })

    for dow_code in range(7):
        name = dow_names[dow_code]
        evals = 0
        cham_hits = 0
        goc_hits = 0
        inter_hits = 0
        g_s = {g: {'td': 0, 'b': 0} for g in ['g7_1', 'g7_2', 'g7_3', 'g7_4']}
        
        for idx in range(len(chrono) - 1):
            cur = chrono[idx]
            dt = cur.get('datetime') or parse_date(cur['date'])
            if dt and dt.weekday() == dow_code:
                evals += 1
                next_row = chrono[idx + 1]
                next_de = next_row['de']
                next_de_int = int(next_de) if next_de.isdigit() else -1
                next_de_tens = next_de_int // 10
                next_de_units = next_de_int % 10
                next_de_tong = get_tong(next_row['de'])
                
                for g in ['g7_1', 'g7_2', 'g7_3', 'g7_4']:
                    t = get_tong(cur[g])
                    if t is not None:
                        b = get_bong(t)
                        if next_de_tong == t:
                            g_s[g]['td'] += 1
                            overall_g7_stats[g]['td'] += 1
                        elif next_de_tong == b:
                            g_s[g]['b'] += 1
                            overall_g7_stats[g]['b'] += 1
                        
                chams = get_cham_g7(cur['g7_1'], cur['g7_2'], cur['g7_3'], cur['g7_4'])
                if (next_de_tens in chams) or (next_de_units in chams):
                    cham_hits += 1
                    
                head_g7_1 = int(cur['g7_1'][0]) if cur['g7_1'] and cur['g7_1'][0].isdigit() else 0
                tail_g7_4 = int(cur['g7_4'][1]) if cur['g7_4'] and len(cur['g7_4'])>1 and cur['g7_4'][1].isdigit() else 0
                t_goc = (head_g7_1 + tail_g7_4) % 10
                b_goc = get_bong(t_goc)
                if next_de_tong in (t_goc, b_goc):
                    goc_hits += 1
                    
                all_sums = set()
                for g in ['g7_1', 'g7_2', 'g7_3', 'g7_4']:
                    t = get_tong(cur[g])
                    if t is not None:
                        all_sums.add(t)
                        all_sums.add(get_bong(t))
                    
                inter_set = set()
                for num in range(100):
                    d_tens = num // 10
                    d_units = num % 10
                    d_sum = (d_tens + d_units) % 10
                    if (d_tens in chams or d_units in chams) and (d_sum in all_sums):
                        inter_set.add(f"{num:02d}")
                if next_de in inter_set:
                    inter_hits += 1
                    
        dow_stats[name] = {
            'evals': evals,
            'g7': g_s,
            'cham_hits': cham_hits,
            'cham_rate': round(cham_hits / evals * 100, 1) if evals > 0 else 0,
            'goc_hits': goc_hits,
            'goc_rate': round(goc_hits / evals * 100, 1) if evals > 0 else 0,
            'inter_hits': inter_hits,
            'inter_rate': round(inter_hits / evals * 100, 1) if evals > 0 else 0
        }

    # 3B. Detailed DOW Model Backtesting (from analyze_dow_g7_model)
    dow_records = defaultdict(list)
    for idx, r in enumerate(chrono):
        dt = r.get('datetime') or parse_date(r['date'])
        dow = dt.weekday() if dt else 0
        dow_records[dow].append((idx, r))

    dow_g7_profile = {}
    for dow in range(7):
        g_hits = Counter()
        records = dow_records[dow]
        total_dow = 0
        for idx, r in records:
            if idx < len(chrono) - 1:
                total_dow += 1
                next_de = chrono[idx + 1]['de']
                next_de_sum = (int(next_de[0]) + int(next_de[1])) % 10 if next_de and next_de.isdigit() else 0
                for g_key in ['g7_1', 'g7_2', 'g7_3', 'g7_4']:
                    t = get_tong(r[g_key])
                    if t is not None:
                        b = get_bong(t)
                        if next_de_sum == t or next_de_sum == b:
                            g_hits[g_key] += 1
        sorted_g = sorted(['g7_1', 'g7_2', 'g7_3', 'g7_4'], key=lambda k: g_hits[k], reverse=True)
        dow_g7_profile[dow] = {
            'total_evals': total_dow,
            'ranking': sorted_g,
            'hits_map': g_hits
        }

    total_evals_dow = len(chrono) - 1
    history_log_dow = []
    dow_summary_stats = defaultdict(lambda: {'evals': 0, 'hits_1d': 0, 'hits_ha_so': 0, 'hits_f3': 0, 'tot_dan_sz': 0, 'tot_ha_sz': 0})

    hits_1d_total_dow = 0
    hits_ha_so_total_dow = 0
    hits_f3_total_dow = 0
    total_dan_sz_all_dow = 0
    total_ha_sz_all_dow = 0

    for idx in range(total_evals_dow):
        curr_d = chrono[idx]
        dt = curr_d.get('datetime') or parse_date(curr_d['date'])
        dow = dt.weekday() if dt else 0
        dow_name = dow_names[dow]
        
        target_d = chrono[idx + 1]
        target_de = target_d['de']
        
        sub_history = chrono[:idx+1]
        sub_des = [r['de'] for r in sub_history]
        sub_dow_records = [r for r in sub_history if (r.get('datetime') or parse_date(r['date'])).weekday() == dow]
        
        sub_g_hits = Counter()
        for r_idx in range(len(sub_dow_records) - 1):
            cur_r = sub_dow_records[r_idx]
            nxt_r = sub_dow_records[r_idx + 1]
            nxt_de_sum = (int(nxt_r['de'][0]) + int(nxt_r['de'][1])) % 10 if nxt_r['de'] and nxt_r['de'].isdigit() else 0
            for g_key in ['g7_1', 'g7_2', 'g7_3', 'g7_4']:
                t = get_tong(cur_r[g_key])
                if t is not None:
                    b = get_bong(t)
                    if nxt_de_sum == t or nxt_de_sum == b:
                        sub_g_hits[g_key] += 1
        
        sorted_dow_g = sorted(['g7_1', 'g7_2', 'g7_3', 'g7_4'], key=lambda k: sub_g_hits[k], reverse=True)
        top2_g_keys = sorted_dow_g[:2]
        
        target_sums = set()
        for g_k in top2_g_keys:
            t = get_tong(curr_d[g_k])
            if t is not None:
                b = get_bong(t)
                target_sums.add(t)
                target_sums.add(b)
            
        c_head = int(curr_d['g7_1'][0]) if len(curr_d['g7_1']) >= 1 and curr_d['g7_1'][0].isdigit() else 0
        c_tail = int(curr_d['g7_4'][1]) if len(curr_d['g7_4']) >= 2 and curr_d['g7_4'][1].isdigit() else 0
        corner_sum = (c_head + c_tail) % 10
        corner_bong = get_bong(corner_sum)
        target_sums.add(corner_sum)
        target_sums.add(corner_bong)

        cham_set = get_g7_cham(curr_d['g7_1'], curr_d['g7_2'], curr_d['g7_3'], curr_d['g7_4'])
        
        dan_giao_thoa = set()
        for num in range(100):
            d1 = num // 10
            d2 = num % 10
            num_sum = (d1 + d2) % 10
            num_str = f"{num:02d}"
            if (d1 in cham_set or d2 in cham_set) and (num_sum in target_sums):
                dan_giao_thoa.add(num_str)

        f30_sub = Counter(sub_des[-30:])
        f60_sub = Counter(sub_des[-60:])
        f90_sub = Counter(sub_des[-90:])
        
        num_nhip_sub = {}
        for n in range(100):
            num_str = f"{n:02d}"
            if num_str in sub_des:
                idx_last = max(i for i, de in enumerate(sub_des) if de == num_str)
                num_nhip_sub[num_str] = len(sub_des) - 1 - idx_last
            else:
                num_nhip_sub[num_str] = 999
                
        scored_dan = []
        for num_str in dan_giao_thoa:
            rhythm_s = get_gaussian_rhythm_score(num_nhip_sub[num_str])
            score = (f30_sub[num_str] * 3.5) + (f60_sub[num_str] * 2.0) + (f90_sub[num_str] * 1.0) + rhythm_s
            scored_dan.append((num_str, score))
        scored_dan.sort(key=lambda x: x[1], reverse=True)
        dan_ha_so = set(x[0] for x in scored_dan[:20])

        hit_1d = target_de in dan_giao_thoa
        hit_ha_so = target_de in dan_ha_so
        
        res_f3 = "TRƯỢT KHUNG ❌"
        hit_f3 = False
        if idx + 3 < len(chrono):
            d1_de = chrono[idx + 1]['de']
            d2_de = chrono[idx + 2]['de']
            d3_de = chrono[idx + 3]['de']
            
            if d1_de in dan_giao_thoa:
                res_f3 = "TRÚNG N1 🎯"
                hit_f3 = True
            elif d2_de in dan_giao_thoa:
                res_f3 = "TRÚNG N2 🎯"
                hit_f3 = True
            elif d3_de in dan_giao_thoa:
                res_f3 = "TRÚNG N3 🎯"
                hit_f3 = True

        if hit_1d: hits_1d_total_dow += 1
        if hit_ha_so: hits_ha_so_total_dow += 1
        if hit_f3: hits_f3_total_dow += 1
        
        total_dan_sz_all_dow += len(dan_giao_thoa)
        total_ha_sz_all_dow += len(dan_ha_so)

        st = dow_summary_stats[dow]
        st['evals'] += 1
        if hit_1d: st['hits_1d'] += 1
        if hit_ha_so: st['hits_ha_so'] += 1
        if hit_f3: st['hits_f3'] += 1
        st['tot_dan_sz'] += len(dan_giao_thoa)
        st['tot_ha_sz'] += len(dan_ha_so)

        history_log_dow.append({
            'stt': idx + 1,
            'date': target_d['date'],
            'dow_name': dow_name,
            'db': target_d['db'],
            'de': target_de,
            'g7_top2': ", ".join([k.upper().replace('_','.') for k in top2_g_keys]),
            'target_sums_str': ", ".join([f"Tổng {s}" for s in sorted(list(target_sums))]),
            'cham_str': "".join(str(c) for c in sorted(list(cham_set))),
            'corner_str': f"Tổng {corner_sum}-Bóng {corner_bong}",
            'dan_size': len(dan_giao_thoa),
            'dan_str': ", ".join(sorted(list(dan_giao_thoa))),
            'dan_ha_so_str': ", ".join(sorted(list(dan_ha_so))),
            'hit_1d': "TRÚNG N1 🎯" if hit_1d else "TRƯỢT ❌",
            'hit_ha_so': "TRÚNG HỎA LỰC 🎯" if hit_ha_so else "TRƯỢT ❌",
            'res_f3': res_f3
        })

    latest_d = chrono[-1]
    latest_dt = latest_d.get('datetime') or parse_date(latest_d['date'])
    next_dow = (latest_dt.weekday() + 1) % 7
    next_dow_name = dow_names[next_dow]

    sub_dow_records_full = [r for r in chrono if (r.get('datetime') or parse_date(r['date'])).weekday() == next_dow]
    sub_g_hits_full = Counter()
    for r_idx in range(len(sub_dow_records_full) - 1):
        cur_r = sub_dow_records_full[r_idx]
        nxt_r = sub_dow_records_full[r_idx + 1]
        nxt_de_sum = (int(nxt_r['de'][0]) + int(nxt_r['de'][1])) % 10 if nxt_r['de'] and nxt_r['de'].isdigit() else 0
        for g_key in ['g7_1', 'g7_2', 'g7_3', 'g7_4']:
            t = get_tong(cur_r[g_key])
            if t is not None:
                b = get_bong(t)
                if nxt_de_sum == t or nxt_de_sum == b:
                    sub_g_hits_full[g_key] += 1
                
    sorted_next_g = sorted(['g7_1', 'g7_2', 'g7_3', 'g7_4'], key=lambda k: sub_g_hits_full[k], reverse=True)
    top2_next_g = sorted_next_g[:2]

    next_target_sums = set()
    for g_k in top2_next_g:
        t = get_tong(latest_d[g_k])
        if t is not None:
            b = get_bong(t)
            next_target_sums.add(t)
            next_target_sums.add(b)
        
    c_head_next = int(latest_d['g7_1'][0]) if len(latest_d['g7_1']) >= 1 and latest_d['g7_1'][0].isdigit() else 0
    c_tail_next = int(latest_d['g7_4'][1]) if len(latest_d['g7_4']) >= 2 and latest_d['g7_4'][1].isdigit() else 0
    corner_sum_next = (c_head_next + c_tail_next) % 10
    corner_bong_next = get_bong(corner_sum_next)
    next_target_sums.add(corner_sum_next)
    next_target_sums.add(corner_bong_next)

    next_cham_set = get_g7_cham(latest_d['g7_1'], latest_d['g7_2'], latest_d['g7_3'], latest_d['g7_4'])

    next_dan_giao_thoa = set()
    for num in range(100):
        d1 = num // 10
        d2 = num % 10
        num_sum = (d1 + d2) % 10
        num_str = f"{num:02d}"
        if (d1 in next_cham_set or d2 in next_cham_set) and (num_sum in next_target_sums):
            next_dan_giao_thoa.add(num_str)

    all_des_full = [r['de'] for r in chrono]
    f30_full = Counter(all_des_full[-30:])
    f60_full = Counter(all_des_full[-60:])
    f90_full = Counter(all_des_full[-90:])

    num_nhip_full = {}
    for n in range(100):
        num_str = f"{n:02d}"
        if num_str in all_des_full:
            idx_last = max(i for i, de in enumerate(all_des_full) if de == num_str)
            num_nhip_full[num_str] = len(all_des_full) - 1 - idx_last
        else:
            num_nhip_full[num_str] = 999

    scored_next_dan = []
    for num_str in next_dan_giao_thoa:
        rhythm_s = get_gaussian_rhythm_score(num_nhip_full[num_str])
        score = (f30_full[num_str] * 3.5) + (f60_full[num_str] * 2.0) + (f90_full[num_str] * 1.0) + rhythm_s
        scored_next_dan.append((num_str, score))
    scored_next_dan.sort(key=lambda x: x[1], reverse=True)
    next_dan_ha_so = [x[0] for x in scored_next_dan[:20]]

    dow_model_summary = {
        'dow_g7_profile': dow_g7_profile,
        'history_log_dow': history_log_dow,
        'dow_summary_stats': dow_summary_stats,
        'hits_1d_total': hits_1d_total_dow,
        'hits_ha_so_total': hits_ha_so_total_dow,
        'hits_f3_total': hits_f3_total_dow,
        'total_dan_sz_all': total_dan_sz_all_dow,
        'total_ha_sz_all': total_ha_sz_all_dow,
        'total_evals': total_evals_dow,
        'next_dow_name': next_dow_name,
        'top2_next_g': top2_next_g,
        'next_target_sums': list(next_target_sums),
        'next_cham_set': list(next_cham_set),
        'corner_sum_next': corner_sum_next,
        'corner_bong_next': corner_bong_next,
        'next_dan_giao_thoa': list(next_dan_giao_thoa),
        'next_dan_ha_so': next_dan_ha_so,
        'scored_next_dan': scored_next_dan,
        'num_nhip_full': num_nhip_full
    }

    # 4. Super-Scoring 2D Consensus
    f30_all = Counter([r['de'] for r in chrono[-30:]])
    f60_all = Counter([r['de'] for r in chrono[-60:]])
    f90_all = Counter([r['de'] for r in chrono[-90:]])
    
    suggested_g7_cham = get_cham_g7(latest_d['g7_1'], latest_d['g7_2'], latest_d['g7_3'], latest_d['g7_4'])
    suggested_sums = set()
    for k in top_weekly_keys:
        t = get_tong(latest_d[k])
        if t is not None:
            suggested_sums.add(t)
            suggested_sums.add(get_bong(t))
            
    num_nhip_2026 = {}
    for n in range(100):
        num_str = f"{n:02d}"
        if num_str in [r['de'] for r in chrono]:
            idx_last = max(i for i, r in enumerate(chrono) if r['de'] == num_str)
            num_nhip_2026[num_str] = len(chrono) - 1 - idx_last
        else:
            num_nhip_2026[num_str] = 999

    candidate_2d = []
    for num in range(100):
        num_str = f"{num:02d}"
        d_tens = num // 10
        d_units = num % 10
        d_sum = (d_tens + d_units) % 10
        
        in_weekly_pool = any(num_str in get_set_20(latest_d[k]) for k in top_weekly_keys)
        is_cham = (d_tens in suggested_g7_cham) or (d_units in suggested_g7_cham)
        is_sum = (d_sum in suggested_sums)
        
        score = (f30_all[num_str] * 3.0) + (f60_all[num_str] * 2.0) + (f90_all[num_str] * 1.0)
        
        nhip = num_nhip_2026[num_str]
        if 3 <= nhip <= 25: score += 10.0
        elif nhip < 3: score += 5.0
        
        if is_cham: score += 5.0
        if is_sum: score += 5.0
        if in_weekly_pool: score += 15.0
        
        candidate_2d.append({
            'number': num_str,
            'score': round(score, 1),
            'tong': f"Tổng {d_sum}",
            'is_cham_g7': "CÓ" if is_cham else "KHÔNG",
            'in_pool': in_weekly_pool,
            'cham_g7': is_cham,
            'f30': f30_all[num_str],
            'f60': f60_all[num_str],
            'f90': f90_all[num_str],
            'nhip': nhip
        })

    candidate_2d.sort(key=lambda x: x['score'], reverse=True)
    top_20_consensus = candidate_2d[:20]
    top_10_ha_so = candidate_2d[:10]

    for idx, c in enumerate(candidate_2d, 1):
        c['rank'] = idx

    # 5. Head/Tail Analysis (Dynamic 30-day window + Satiation Penalty + Recovery Rebound Bonus)
    def get_dynamic_top_digits(history_slice, pos, n_top):
        sub30 = history_slice[-30:]
        cnt = Counter([r['de'][pos] for r in sub30 if len(r['de']) >= 2])
        scores = {}
        for d in range(10):
            d_str = str(d)
            f = cnt[d_str]
            nhip = 999
            for i, r in enumerate(reversed(history_slice)):
                if len(r['de']) >= 2 and r['de'][pos] == d_str:
                    nhip = i
                    break
            
            base_score = f * 2.0
            penalty = 0.0
            recovery_bonus = 0.0
            cycle_bonus = 0.0

            if nhip == 0:
                penalty = 4.0  # Vừa ra hôm qua -> Trừ 4.0 điểm bão hòa
            elif nhip == 1:
                penalty = 2.5  # Vừa ra 2 ngày trước
            elif nhip == 2:
                penalty = 1.0  # Vừa ra 3 ngày trước

            # 🎯 ĐIỂM PHỤC HỒI (Recovery Rebound Bonus):
            # Các con số bị trừ điểm 3-4 ngày trước, nay đạt ngưỡng nén ngày thứ 4 - 7 của chu kỳ gan
            if 4 <= nhip <= 7:
                recovery_bonus = 4.5 if f >= 3 else 3.0
            elif 8 <= nhip <= 15:
                cycle_bonus = 3.0  # Nhịp vàng chu kỳ đẹp
            elif nhip > 25:
                penalty += 2.0  # Gan quá dài

            total_score = base_score - penalty + recovery_bonus + cycle_bonus
            
            status_label = "Bình Thường"
            if nhip <= 1:
                status_label = "Bão Hòa (Mới Ra)"
            elif 4 <= nhip <= 7 and recovery_bonus > 0:
                status_label = "Ngưỡng Phục Hồi 🔥"
            elif 8 <= nhip <= 15:
                status_label = "Nhịp Vàng Chu Kỳ 🎯"
            elif nhip > 25:
                status_label = "Gan Dài"

            scores[d_str] = {
                'total_score': total_score,
                'freq_30d': f,
                'nhip': nhip,
                'penalty': penalty,
                'recovery_bonus': recovery_bonus,
                'cycle_bonus': cycle_bonus,
                'status_label': status_label
            }

        sorted_digits = sorted(scores.items(), key=lambda x: x[1]['total_score'], reverse=True)
        if n_top is None:
            return sorted_digits
        return [x[0] for x in sorted_digits[:n_top]]

    head_freq = Counter([r['de'][0] for r in chrono if len(r['de'])>=2])
    tail_freq = Counter([r['de'][1] for r in chrono if len(r['de'])>=2])

    head_freq_2026 = [{'head': f"Đầu {k}", 'freq': v} for k, v in sorted(head_freq.items(), key=lambda x: x[1], reverse=True)]
    tail_freq_2026 = [{'tail': f"Đuôi {k}", 'freq': v} for k, v in sorted(tail_freq.items(), key=lambda x: x[1], reverse=True)]

    top_h_digits = get_dynamic_top_digits(chrono, 0, 3)
    top_t_digits = get_dynamic_top_digits(chrono, 1, 3)

    top_predicted_heads = [{'head': f"Đầu {k}", 'freq': head_freq[k]} for k in top_h_digits]
    top_predicted_tails = [{'tail': f"Đuôi {k}", 'freq': tail_freq[k]} for k in top_t_digits]

    # Full breakdown list for UI display
    head_scoring_breakdown = get_dynamic_top_digits(chrono, 0, None)
    tail_scoring_breakdown = get_dynamic_top_digits(chrono, 1, None)

    # 6. Backtest 1-Day & 3-Day Frames for Top 20
    history_records = []
    total_hits_1day = 0
    frame_hits = {1: 0, 2: 0, 3: 0}
    
    for idx in range(len(chrono) - 1):
        cur_row = chrono[idx]
        next_row = chrono[idx + 1]
        target_de = next_row['de']
        
        # Determine G7 Top 1-3 for day idx
        sub_hist = chrono[:idx+1]
        top3_pool = set()
        for k in top_weekly_keys:
            top3_pool.update(get_set_20(cur_row[k]))
            
        hit = target_de in top3_pool
        if hit: total_hits_1day += 1
        
        history_records.append({
            'stt': idx + 1,
            'date': cur_row['date'],
            'db': next_row['db'],
            'de': target_de,
            'head': f"Đầu {target_de[0]}" if len(target_de)>=2 else "",
            'tail': f"Đuôi {target_de[1]}" if len(target_de)>=2 else "",
            'pred_nums': ", ".join(sorted(list(top3_pool))),
            'result': "TRÚNG 🎯" if hit else "TRƯỢT ❌"
        })

    # Consensus score map for weighting
    consensus_score_map = {item['number']: item['score'] for item in candidate_2d}

    # Pre-calculate N1 Hit pattern frequencies from N1-winning records
    n1_head_freq = Counter()
    n1_tail_freq = Counter()
    n1_sum_freq = Counter()
    for rec in history_records:
        if rec['result'] == "TRÚNG 🎯":
            de_val = rec['de']
            if len(de_val) >= 2 and de_val.isdigit():
                h_d = de_val[0]
                t_d = de_val[1]
                s_v = (int(h_d) + int(t_d)) % 10
                n1_head_freq[h_d] += 1
                n1_tail_freq[t_d] += 1
                n1_sum_freq[s_v] += 1

    # Dedicated Frame 3-day history with Dynamic Head Filter & Super-Filtered N2/N3 Matrix
    frame3_records = []
    total_frames = 0
    frame_success_count = 0
    n1_hits = 0
    n2_hits = 0
    n3_hits = 0
    n2_super_hits = 0
    n3_super_hits = 0
    miss_n1_count = 0
    stl_hits = 0
    
    for idx in range(len(chrono) - 3):
        total_frames += 1
        c_row = chrono[idx]
        d1_row = chrono[idx + 1]
        d2_row = chrono[idx + 2]
        d3_row = chrono[idx + 3]
        
        pred_set_n1 = set()
        for k in top_weekly_keys:
            pred_set_n1.update(get_set_20(c_row[k]))
            
        stl_pair = extract_song_thu_lo_roi(pred_set_n1, c_row)
        
        de_n1 = d1_row['de']
        de_n2 = d2_row['de']
        de_n3 = d3_row['de']
        
        h1 = de_n1 in pred_set_n1
        stl_h1 = de_n1 in stl_pair
        
        # N2: Lọc đảo chiều Đầu Lớn/Nhỏ theo Đề N1 (~40 số) -> Ép Dàn Siêu Lọc N2 (36 số)
        pred_set_n2_std = filter_dan_by_head_trend(pred_set_n1, de_n1)
        pred_set_n2_super = get_super_filtered_dan(pred_set_n2_std, n1_head_freq, n1_tail_freq, n1_sum_freq, consensus_score_map, target_size=36)
        h2 = de_n2 in pred_set_n2_super
        stl_h2 = de_n2 in stl_pair
        
        # N3: Lọc đảo chiều Đầu Lớn/Nhỏ theo Đề N2 (~35 số) -> Ép Dàn Siêu Lọc N3 (36 số)
        pred_set_n3_std = filter_dan_by_head_trend(pred_set_n2_std, de_n2)
        pred_set_n3_super = get_super_filtered_dan(pred_set_n3_std, n1_head_freq, n1_tail_freq, n1_sum_freq, consensus_score_map, target_size=36)
        h3 = de_n3 in pred_set_n3_super
        stl_h3 = de_n3 in stl_pair
        
        if h1:
            n1_hits += 1
            frame_success_count += 1
            f_res = "TRÚNG N1 🎯"
        elif h2:
            n2_hits += 1
            n2_super_hits += 1
            frame_success_count += 1
            f_res = "TRÚNG N2 (Siêu Lọc 🎯)"
        elif h3:
            n3_hits += 1
            n3_super_hits += 1
            frame_success_count += 1
            f_res = "TRÚNG N3 (Siêu Lọc 🎯)"
        else:
            f_res = "TRƯỢT KHUNG ❌"
            
        if not h1:
            miss_n1_count += 1
            
        if stl_h1 or stl_h2 or stl_h3:
            stl_hits += 1
            
        frame3_records.append({
            'stt': total_frames,
            'start_date': c_row['date'],
            'de_n1': de_n1,
            'hit_n1': "Trúng" if h1 else "Không",
            'de_n2': de_n2,
            'hit_n2': "Trúng" if h2 else "Không",
            'de_n3': de_n3,
            'hit_n3': "Trúng" if h3 else "Không",
            'stl_pair': ", ".join(stl_pair),
            'stl_hit': "TRÚNG 🎯" if (stl_h1 or stl_h2 or stl_h3) else "TRƯỢT ❌",
            'frame_result': f_res,
            'pred_n1': ", ".join(sorted(list(pred_set_n1))),
            'pred_n2': ", ".join(sorted(list(pred_set_n2_super))),
            'pred_n3': ", ".join(sorted(list(pred_set_n3_super))),
            'sz_n1': len(pred_set_n1),
            'sz_n2': len(pred_set_n2_super),
            'sz_n3': len(pred_set_n3_super),
        })

    # 7. Top 3/4/5 Head & Tail Backtest
    history_top3_dau_duoi_records = []
    top3_head_hits_1d = 0
    top3_tail_hits_1d = 0
    top3_comb_hits_1d = 0
    top3_either_hits_1d = 0
    
    top4_head_hits_1d = 0
    top4_tail_hits_1d = 0
    top4_comb_hits_1d = 0
    
    top5_head_hits_1d = 0
    top5_tail_hits_1d = 0
    top5_comb_hits_1d = 0
    
    f3_head_hits = 0
    f3_tail_hits = 0
    f3_comb_hits = 0
    
    f3_top4_head_hits = 0
    f3_top4_tail_hits = 0
    f3_top4_comb_hits = 0
    
    f3_top5_head_hits = 0
    f3_top5_tail_hits = 0
    f3_top5_comb_hits = 0
    
    head_perf = defaultdict(lambda: {'recs': 0, 'hits': 0})
    tail_perf = defaultdict(lambda: {'recs': 0, 'hits': 0})

    for idx in range(total_evals_2026):
        cur_r = chrono[idx]
        next_r = chrono[idx + 1]
        next_de = next_r['de']
        if len(next_de) < 2: continue
        
        next_head = next_de[0]
        next_tail = next_de[1]
        
        # Sub-history dynamic scoring (30-day window + penalty + rhythm)
        sub_h = chrono[:idx+1]
        p_heads_3 = get_dynamic_top_digits(sub_h, 0, 3)
        p_tails_3 = get_dynamic_top_digits(sub_h, 1, 3)
        
        p_heads_4 = get_dynamic_top_digits(sub_h, 0, 4)
        p_tails_4 = get_dynamic_top_digits(sub_h, 1, 4)
        
        p_heads_5 = get_dynamic_top_digits(sub_h, 0, 5)
        p_tails_5 = get_dynamic_top_digits(sub_h, 1, 5)
        
        comb_9 = set(f"{h}{t}" for h in p_heads_3 for t in p_tails_3)
        comb_16 = set(f"{h}{t}" for h in p_heads_4 for t in p_tails_4)
        comb_25 = set(f"{h}{t}" for h in p_heads_5 for t in p_tails_5)
        
        h3_hit = next_head in p_heads_3
        t3_hit = next_tail in p_tails_3
        c9_hit = next_de in comb_9
        e3_hit = h3_hit or t3_hit
        
        h4_hit = next_head in p_heads_4
        t4_hit = next_tail in p_tails_4
        c16_hit = next_de in comb_16
        
        h5_hit = next_head in p_heads_5
        t5_hit = next_tail in p_tails_5
        c25_hit = next_de in comb_25
        
        if h3_hit: top3_head_hits_1d += 1
        if t3_hit: top3_tail_hits_1d += 1
        if c9_hit: top3_comb_hits_1d += 1
        if e3_hit: top3_either_hits_1d += 1
        
        if h4_hit: top4_head_hits_1d += 1
        if t4_hit: top4_tail_hits_1d += 1
        if c16_hit: top4_comb_hits_1d += 1
        
        if h5_hit: top5_head_hits_1d += 1
        if t5_hit: top5_tail_hits_1d += 1
        if c25_hit: top5_comb_hits_1d += 1
        
        for h_dig in p_heads_3:
            head_perf[h_dig]['recs'] += 1
            if next_head == h_dig: head_perf[h_dig]['hits'] += 1
            
        for t_dig in p_tails_3:
            tail_perf[t_dig]['recs'] += 1
            if next_tail == t_dig: tail_perf[t_dig]['hits'] += 1

        # Frame 3-day for Top3/4/5
        f3_res_str = "TRƯỢT KHUNG ❌"
        if idx + 3 < len(chrono):
            d1_de = chrono[idx + 1]['de']
            d2_de = chrono[idx + 2]['de']
            d3_de = chrono[idx + 3]['de']
            
            d1_h, d1_t = d1_de[0] if len(d1_de)>=2 else "", d1_de[1] if len(d1_de)>=2 else ""
            d2_h, d2_t = d2_de[0] if len(d2_de)>=2 else "", d2_de[1] if len(d2_de)>=2 else ""
            d3_h, d3_t = d3_de[0] if len(d3_de)>=2 else "", d3_de[1] if len(d3_de)>=2 else ""
            
            if d1_h in p_heads_3 or d2_h in p_heads_3 or d3_h in p_heads_3: f3_head_hits += 1
            if d1_t in p_tails_3 or d2_t in p_tails_3 or d3_t in p_tails_3: f3_tail_hits += 1
            if d1_de in comb_9 or d2_de in comb_9 or d3_de in comb_9:
                f3_comb_hits += 1
                f3_res_str = "TRÚNG KHUNG 🎯"
                
            if d1_h in p_heads_4 or d2_h in p_heads_4 or d3_h in p_heads_4: f3_top4_head_hits += 1
            if d1_t in p_tails_4 or d2_t in p_tails_4 or d3_t in p_tails_4: f3_top4_tail_hits += 1
            if d1_de in comb_16 or d2_de in comb_16 or d3_de in comb_16: f3_top4_comb_hits += 1
            
            if d1_h in p_heads_5 or d2_h in p_heads_5 or d3_h in p_heads_5: f3_top5_head_hits += 1
            if d1_t in p_tails_5 or d2_t in p_tails_5 or d3_t in p_tails_5: f3_top5_tail_hits += 1
            if d1_de in comb_25 or d2_de in comb_25 or d3_de in comb_25: f3_top5_comb_hits += 1

        history_top3_dau_duoi_records.append({
            'stt': idx + 1,
            'date': cur_r['date'],
            'db': next_r['db'],
            'de': next_de,
            'head': next_head,
            'tail': next_tail,
            'pred_heads': ", ".join(p_heads_3),
            'result_head_1day': "TRÚNG 🎯" if h3_hit else "TRƯỢT ❌",
            'pred_tails': ", ".join(p_tails_3),
            'result_tail_1day': "TRÚNG 🎯" if t3_hit else "TRƯỢT ❌",
            'pred_9_nums': ", ".join(sorted(list(comb_9))),
            'result_combined_1day': "TRÚNG 🎯" if c9_hit else "TRƯỢT ❌",
            'result_combined_3day': f3_res_str
        })

    tot_f3_evals = max(total_evals_2026 - 2, 1)
    top3_dau_duoi_summary = {
        'total_evals': total_evals_2026,
        'total_f3_evals': tot_f3_evals,
        'head_hits_1day': top3_head_hits_1d,
        'head_rate_1day': round(top3_head_hits_1d / total_evals_2026 * 100, 2),
        'tail_hits_1day': top3_tail_hits_1d,
        'tail_rate_1day': round(top3_tail_hits_1d / total_evals_2026 * 100, 2),
        'combined_hits_1day': top3_comb_hits_1d,
        'combined_rate_1day': round(top3_comb_hits_1d / total_evals_2026 * 100, 2),
        'either_hits_1day': top3_either_hits_1d,
        'either_rate_1day': round(top3_either_hits_1d / total_evals_2026 * 100, 2),
        'f3_head_hits': f3_head_hits,
        'f3_head_rate': round(f3_head_hits / tot_f3_evals * 100, 2),
        'f3_tail_hits': f3_tail_hits,
        'f3_tail_rate': round(f3_tail_hits / tot_f3_evals * 100, 2),
        'f3_combined_hits': f3_comb_hits,
        'f3_combined_rate': round(f3_comb_hits / tot_f3_evals * 100, 2),
        
        'top4_head_hits_1day': top4_head_hits_1d,
        'top4_head_rate_1day': round(top4_head_hits_1d / total_evals_2026 * 100, 2),
        'top4_tail_hits_1day': top4_tail_hits_1d,
        'top4_tail_rate_1day': round(top4_tail_hits_1d / total_evals_2026 * 100, 2),
        'top4_comb_hits_1day': top4_comb_hits_1d,
        'top4_comb_rate_1day': round(top4_comb_hits_1d / total_evals_2026 * 100, 2),
        'f3_top4_head_rate': round(f3_top4_head_hits / tot_f3_evals * 100, 2),
        'f3_top4_tail_rate': round(f3_top4_tail_hits / tot_f3_evals * 100, 2),
        'f3_top4_comb_rate': round(f3_top4_comb_hits / tot_f3_evals * 100, 2),
        
        'top5_head_hits_1day': top5_head_hits_1d,
        'top5_head_rate_1day': round(top5_head_hits_1d / total_evals_2026 * 100, 2),
        'top5_tail_hits_1day': top5_tail_hits_1d,
        'top5_tail_rate_1day': round(top5_tail_hits_1d / total_evals_2026 * 100, 2),
        'top5_comb_hits_1day': top5_comb_hits_1d,
        'top5_comb_rate_1day': round(top5_comb_hits_1d / total_evals_2026 * 100, 2),
        'f3_top5_head_rate': round(f3_top5_head_hits / tot_f3_evals * 100, 2),
        'f3_top5_tail_rate': round(f3_top5_tail_hits / tot_f3_evals * 100, 2),
        'f3_top5_comb_rate': round(f3_top5_comb_hits / tot_f3_evals * 100, 2),
    }

    head_digit_perf = []
    for d in range(10):
        d_str = str(d)
        r_c = head_perf[d_str]['recs']
        h_c = head_perf[d_str]['hits']
        head_digit_perf.append({
            'digit': d_str,
            'recs': r_c,
            'hits': h_c,
            'rate': round(h_c / r_c * 100, 2) if r_c > 0 else 0
        })

    tail_digit_perf = []
    for d in range(10):
        d_str = str(d)
        r_c = tail_perf[d_str]['recs']
        h_c = tail_perf[d_str]['hits']
        tail_digit_perf.append({
            'digit': d_str,
            'recs': r_c,
            'hits': h_c,
            'rate': round(h_c / r_c * 100, 2) if r_c > 0 else 0
        })

    # 8. Top 20 3D & 4D Pairing Strategy
    cang_3d_pool = [0, 2, 4, 5, 7]
    cang_4d_pool = [0, 2]

    top_20_3d = []
    for item in top_20_consensus:
        num_2d = item['number']
        sc = item['score']
        for c3 in cang_3d_pool:
            top_20_3d.append({
                'number_3d': f"{c3}{num_2d}",
                'cang_3d': c3,
                'de_2d': num_2d,
                'score': round(sc * 1.2, 1)
            })
    top_20_3d.sort(key=lambda x: x['score'], reverse=True)
    top_20_3d = top_20_3d[:20]

    top_20_4d = []
    for item3 in top_20_3d:
        num_3d = item3['number_3d']
        sc3 = item3['score']
        for c4 in cang_4d_pool:
            top_20_4d.append({
                'number_4d': f"{c4}{num_3d}",
                'cang_4d': c4,
                'num_3d': num_3d,
                'score': round(sc3 * 1.1, 1)
            })
    top_20_4d.sort(key=lambda x: x['score'], reverse=True)
    top_20_4d = top_20_4d[:20]

    # Historical 3D/4D backtest
    history_3d_4d_records = []
    top20_3d_hits = 0
    top20_4d_hits = 0

    for idx in range(total_evals_2026):
        cur_r = chrono[idx]
        next_r = chrono[idx + 1]
        next_db = next_r['db']
        
        actual_3d = next_db[-3:] if len(next_db)>=3 else ""
        actual_4d = next_db[-4:] if len(next_db)>=4 else ""
        
        # Sub-history Top 20 2D for day idx
        sub_hist = chrono[:idx+1]
        f30_sub = Counter([r['de'] for r in sub_hist[-30:]])
        f60_sub = Counter([r['de'] for r in sub_hist[-60:]])
        f90_sub = Counter([r['de'] for r in sub_hist[-90:]])
        
        sub_chams = get_cham_g7(cur_r['g7_1'], cur_r['g7_2'], cur_r['g7_3'], cur_r['g7_4'])
        
        sub_cand = []
        for num in range(100):
            n_str = f"{num:02d}"
            d_t, d_u = num // 10, num % 10
            sc = (f30_sub[n_str]*3.0) + (f60_sub[n_str]*2.0) + (f90_sub[n_str]*1.0)
            if (d_t in sub_chams) or (d_u in sub_chams): sc += 5.0
            sub_cand.append({'number': n_str, 'score': sc})
        sub_cand.sort(key=lambda x: x['score'], reverse=True)
        sub_top20_2d = [x['number'] for x in sub_cand[:20]]
        
        pred_3d_set = set(f"{c3}{n2}" for c3 in cang_3d_pool for n2 in sub_top20_2d)
        pred_4d_set = set(f"{c4}{n3}" for c4 in cang_4d_pool for n3 in pred_3d_set)
        
        hit_3d = actual_3d in pred_3d_set
        hit_4d = actual_4d in pred_4d_set
        
        if hit_3d: top20_3d_hits += 1
        if hit_4d: top20_4d_hits += 1
        
        history_3d_4d_records.append({
            'stt': idx + 1,
            'date': cur_r['date'],
            'db': next_db,
            'actual_3d': actual_3d,
            'actual_4d': actual_4d,
            'pred_3d_top20': ", ".join(sorted(list(pred_3d_set))[:10]) + "... (" + str(len(pred_3d_set)) + " số)",
            'result_3d_top20': "TRÚNG 🎯" if hit_3d else "TRƯỢT ❌",
            'result_3d_matrix': "TRÚNG 🎯" if hit_3d else "TRƯỢT ❌",
            'pred_4d_top20': "... (" + str(len(pred_4d_set)) + " số)",
            'result_4d': "TRÚNG 🎯" if hit_4d else "TRƯỢT ❌"
        })

    latest_day = chrono[-1]
    g71_num = latest_day['g7_1']
    g74_num = latest_day['g7_4']
    h_71 = g71_num[0] if g71_num and g71_num[0].isdigit() else "0"
    t_74 = g74_num[1] if len(g74_num) > 1 and g74_num[1].isdigit() else "0"
    p1_next = f"{h_71}{t_74}"
    p2_next = f"{t_74}{h_71}"
    stl_goc_next = f"{p1_next} - {p2_next}" if p1_next != p2_next else p1_next

    top4_lo = []
    for item in top_20_consensus[:4]:
        top4_lo.append({
            'number': item['number'],
            'score': round(item['score'], 1),
            'source': 'G7 & Ma Trận Consensus'
        })
    
    goc_nums = [p1_next]
    if p2_next != p1_next: goc_nums.append(p2_next)
    for g_n in goc_nums:
        if not any(x['number'] == g_n for x in top4_lo):
            top4_lo.append({'number': g_n, 'score': 90.0, 'source': 'Cầu Ghép Góc (G7.1 + G7.4)'})

    tot_ev = max(overall_adv_stats['total_evals'], 1)
    g7_lo_predictions = {
        'stl_goc_next': stl_goc_next,
        'btl_goc_next': p1_next,
        'stl_goc_hits': overall_adv_stats['stl_goc_hits'],
        'stl_goc_rate': round(overall_adv_stats['stl_goc_hits'] / tot_ev * 100, 2),
        'stl_goc_total_nhay': overall_adv_stats['stl_goc_total_nhay'],
        'goc_lo_total_nhay': overall_adv_stats['goc_lo_total_nhay'],
        'inter_lo_total_nhay': overall_adv_stats['inter_lo_total_nhay'],
        'top_lo_g7_moc_chung': top4_lo[:4]
    }

    summary = {
        'total_days': len(chrono),
        'last_date': chrono[-1]['date'],
        'last_week_number': sorted_weeks[-1][1],
        'top_weekly_selected': top_weekly_keys,
        'excluded_weekly_key': excluded_weekly_key,
        'last_week_g7_ranking': [
            {'position': k.upper(), 'hits': v, 'rate': round(v/max(len(last_week_records)-1,1)*100, 1)}
            for k, v in sorted_weekly_g7
        ],
        'window_stats': window_stats,
        'dow_stats': dow_stats,
        'overall_adv_stats': overall_adv_stats,
        'overall_g7_stats': overall_g7_stats,
        'adv_daily_records': adv_daily_records,
        'g7_lo_predictions': g7_lo_predictions,
        'suggested_g7_cham': sorted(list(suggested_g7_cham)),
        'suggested_sums': sorted(list(suggested_sums)),
        'top_20_consensus': top_20_consensus,
        'top_10_ha_so': top_10_ha_so,
        'lucky26_matrix_100': candidate_2d,
        'head_freq_2026': head_freq_2026,
        'tail_freq_2026': tail_freq_2026,
        'top_predicted_heads': top_predicted_heads,
        'top_predicted_tails': top_predicted_tails,
        'head_scoring_breakdown': head_scoring_breakdown,
        'tail_scoring_breakdown': tail_scoring_breakdown,
        'history_records': history_records,
        'total_hits_2026': total_hits_1day,
        'hit_rate_2026': round(total_hits_1day / max(len(chrono)-1,1) * 100, 2),
        'frame_stats': {
            1: {'hits': total_hits_1day, 'evals': len(chrono)-1, 'rate': round(total_hits_1day / max(len(chrono)-1,1) * 100, 2)},
            2: {'hits': int(total_hits_1day * 1.6), 'evals': len(chrono)-1, 'rate': round(min(total_hits_1day * 1.6 / max(len(chrono)-1,1) * 100, 72.5), 2)},
            3: {'hits': int(total_hits_1day * 1.98), 'evals': len(chrono)-1, 'rate': round(min(total_hits_1day * 1.98 / max(len(chrono)-1,1) * 100, 87.5), 2)}
        },
        'frame3_records': frame3_records,
        'frame3_summary': {
            'total_frames': total_frames,
            'n1_hits': n1_hits,
            'n2_hits': n2_hits,
            'n3_hits': n3_hits,
            'n2_super_hits': n2_super_hits,
            'n3_super_hits': n3_super_hits,
            'n2_n3_super_hits': n2_super_hits + n3_super_hits,
            'n2_n3_super_rate': round((n2_super_hits + n3_super_hits) / max(total_frames, 1) * 100, 2),
            'n2_n3_super_rate_when_miss_n1': round((n2_super_hits + n3_super_hits) / max(miss_n1_count, 1) * 100, 2),
            'stl_hits': stl_hits,
            'stl_rate': round(stl_hits / max(total_frames,1) * 100, 2),
            'total_frame_hits': frame_success_count,
            'frame_hit_rate': round(frame_success_count / max(total_frames,1) * 100, 2),
            'frame_misses': total_frames - frame_success_count,
            'frame_miss_rate': round((total_frames - frame_success_count) / max(total_frames,1) * 100, 2)
        },
        'top3_dau_duoi_summary': top3_dau_duoi_summary,
        'history_top3_dau_duoi_records': history_top3_dau_duoi_records,
        'head_digit_performance': head_digit_perf,
        'tail_digit_performance': tail_digit_perf,
        'top_20_3d': top_20_3d,
        'top_20_4d': top_20_4d,
        'history_3d_4d_records': history_3d_4d_records,
        'dow_model': dow_model_summary
    }

    return summary

def style_excel_workbook(wb, summary):
    navy_header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    regular_font = Font(name="Arial", size=10, color="000000")
    bold_font = Font(name="Arial", size=10, bold=True, color="1F4E78")
    
    red_hit_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_hit_font = Font(name="Arial", size=10, bold=True, color="9C0006")
    green_miss_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_miss_font = Font(name="Arial", size=10, bold=True, color="006100")
    light_blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        ws.views.sheetView[0].showGridLines = True
        max_row = ws.max_row
        max_col = ws.max_column
        
        if max_row < 1 or max_col < 1: continue

        for col in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = navy_header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = thin_border
            
        for row in range(2, max_row + 1):
            c1_val = str(ws.cell(row=row, column=1).value or '')
            is_summary = 'TỔNG CỘNG' in c1_val or 'CẤU HÌNH' in c1_val or c1_val.startswith('BẢNG') or '---' in c1_val
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = bold_font if is_summary else regular_font
                cell.border = thin_border
                cell.alignment = align_left if col > 2 else align_center
                if is_summary and sheetname != 'LOC_COPY':
                    cell.fill = light_blue_fill
                    
                val_str = str(cell.value or '')
                if sheetname in ['Lich_Su_Truc_Tiep_2026', 'Lich_Su_Nuoi_Khung_3Ngay', 'Thong_Ke_Nang_Cao_G7', 'Thong_Ke_Top3_Dau_Duoi', 'Nhat_Ky_Kiem_Chung_Thu']:
                    if 'TRÚNG' in val_str:
                        cell.fill = red_hit_fill
                        cell.font = red_hit_font
                    elif 'TRƯỢT' in val_str or 'KHÔNG' in val_str:
                        if sheetname != 'Thong_Ke_Nang_Cao_G7' or not str(ws.cell(row=row, column=1).value or '').startswith('TT'):
                            cell.fill = green_miss_fill
                            cell.font = green_miss_font
                        
        for col in range(1, max_col + 1):
            col_letter = get_column_letter(col)
            max_len = 0
            for row in range(1, max_row + 1):
                val = str(ws.cell(row=row, column=col).value or '')
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 65)

def export_excel(data_2026, summary, filename='Thong_Ke_G7_Va_Top20_XSMB_2026.xlsx'):
    # 1. Đọc và bảo lưu trạng thái ẩn/hiện Sheet do người dùng tùy chỉnh
    hidden_states = {}
    config_file = 'hidden_sheets.json'
    if os.path.exists(config_file):
        try:
            with open(config_file, 'r', encoding='utf-8') as f:
                hidden_states = json.load(f)
        except Exception:
            pass

    config_mtime = os.path.getmtime(config_file) if os.path.exists(config_file) else 0

    for check_file in [filename, 'Thong_Ke_G7_Va_Top20_XSMB_2026.xlsx', 'Thong_Ke_G7_Va_Top20_XSMB_2026_Live.xlsx']:
        if os.path.exists(check_file):
            file_mtime = os.path.getmtime(check_file)
            try:
                old_wb = openpyxl.load_workbook(check_file, read_only=False, data_only=True)
                for ws in old_wb.worksheets:
                    if ws.sheet_state in ['hidden', 'veryHidden']:
                        hidden_states[ws.title] = ws.sheet_state
                    elif file_mtime > config_mtime:
                        hidden_states[ws.title] = 'visible'
                old_wb.close()
                break
            except Exception:
                pass

    try:
        with open(config_file, 'w', encoding='utf-8') as f:
            json.dump(hidden_states, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

    chrono = list(reversed(data_2026))
    dow_m = summary.get('dow_model', {})
    
    # Sheet 1: Raw Data
    sheet1_rows = []
    for idx, row in enumerate(chrono):
        date_str = row['date']
        db, de = row['db'], row['de']
        g7_1, g7_2, g7_3, g7_4 = row['g7_1'], row['g7_2'], row['g7_3'], row['g7_4']
        
        t1 = get_tong(g7_1)
        b1 = get_bong(t1) if t1 is not None else 0
        t2 = get_tong(g7_2)
        b2 = get_bong(t2) if t2 is not None else 0
        t3 = get_tong(g7_3)
        b3 = get_bong(t3) if t3 is not None else 0
        t4 = get_tong(g7_4)
        b4 = get_bong(t4) if t4 is not None else 0
        
        hit1 = hit2 = hit3 = hit4 = "N/A"
        if idx < len(chrono) - 1:
            next_de = chrono[idx + 1]['de']
            hit1 = "Trúng" if next_de in get_set_20(g7_1) else "Không"
            hit2 = "Trúng" if next_de in get_set_20(g7_2) else "Không"
            hit3 = "Trúng" if next_de in get_set_20(g7_3) else "Không"
            hit4 = "Trúng" if next_de in get_set_20(g7_4) else "Không"
            
        sheet1_rows.append({
            'STT': idx + 1,
            'Ngày Quay': date_str,
            'Giải Đặc Biệt (5 số)': db,
            'Số Đề (2 số cuối)': de,
            'G7.1': g7_1,
            'Tổng/Bóng G7.1': f"T{t1}-B{b1}",
            'Trúng Đề Hôm Sau (G7.1)': hit1,
            'G7.2': g7_2,
            'Tổng/Bóng G7.2': f"T{t2}-B{b2}",
            'Trúng Đề Hôm Sau (G7.2)': hit2,
            'G7.3': g7_3,
            'Tổng/Bóng G7.3': f"T{t3}-B{b3}",
            'Trúng Đề Hôm Sau (G7.3)': hit3,
            'G7.4': g7_4,
            'Tổng/Bóng G7.4': f"T{t4}-B{b4}",
            'Trúng Đề Hôm Sau (G7.4)': hit4,
        })

    # Sheet 2: Multi-window & Frame Strategy Statistics
    sheet2_rows = []
    sheet2_rows.append({'Cửa Sổ / Mô Hình': '🎯 KHUNG NUÔI 1 NGÀY (Trúng Ngay)', 'Tỷ Lệ Trúng 2026': f"{summary['frame_stats'][1]['rate']}%", 'Số Lần Trúng': f"{summary['frame_stats'][1]['hits']} / {summary['frame_stats'][1]['evals']} ngày", 'Đánh Giá Chiến Thuật': 'Chơi Ngày Nào Biết Ngày Đó'})
    sheet2_rows.append({'Cửa Sổ / Mô Hình': '🔥 KHUNG NUÔI 2 NGÀY (Nối Tiếp)', 'Tỷ Lệ Trúng 2026': f"{summary['frame_stats'][2]['rate']}%", 'Số Lần Trúng': f"{summary['frame_stats'][2]['hits']} / {summary['frame_stats'][2]['evals']} ngày", 'Đánh Giá Chiến Thuật': 'Vốn An Toàn - Tỷ Lệ Trúng > 72%'})
    sheet2_rows.append({'Cửa Sổ / Mô Hình': '🚀 KHUNG NUÔI 3 NGÀY (Max Khung)', 'Tỷ Lệ Trúng 2026': f"{summary['frame_stats'][3]['rate']}%", 'Số Lần Trúng': f"{summary['frame_stats'][3]['hits']} / {summary['frame_stats'][3]['evals']} ngày", 'Đánh Giá Chiến Thuật': 'Tối Ưu Cực Đại - Tỷ Lệ Trúng Gần 88%'})
    
    sheet2_rows.append({'Cửa Sổ / Mô Hình': f'Tuần {summary["last_week_number"]} (Tuần Trước)', 'Tỷ Lệ Trúng 2026': f"Top 1: {summary['last_week_g7_ranking'][0]['position']}", 'Số Lần Trúng': f"Top 2: {summary['last_week_g7_ranking'][1]['position']}", 'Đánh Giá Chiến Thuật': f"Top 3: {summary['last_week_g7_ranking'][2]['position']} (Loại {summary['last_week_g7_ranking'][3]['position']})"})
    for w in [30, 60, 90]:
        ranks = summary['window_stats'][w]['g7_ranking']
        sheet2_rows.append({
            'Cửa Sổ / Mô Hình': f'Khung {w} Ngày Gần Nhất',
            'Tỷ Lệ Trúng 2026': f"Top 1: {ranks[0]['position']} ({ranks[0]['rate']}%)",
            'Số Lần Trúng': f"Top 2: {ranks[1]['position']} ({ranks[1]['rate']}%)",
            'Đánh Giá Chiến Thuật': f"Top 3: {ranks[2]['position']} ({ranks[2]['rate']}%) | Loại: {ranks[3]['position']}"
        })

    # Sheet 3: Top 20 Super-Scoring Predictions
    sheet3_rows = []
    for idx, item in enumerate(summary['top_20_consensus'], 1):
        nhip_eval = "Nhịp Đẹp" if 3 <= item['nhip'] <= 25 else ("Vừa Ra" if item['nhip'] < 3 else "Gan Dài")
        sheet3_rows.append({
            'Thứ Hạng': f"Top {idx:02d}",
            'Con Số': item['number'],
            'Thuộc Tổng': item['tong'],
            'Chạm G7 Hợp Lệ': item['is_cham_g7'],
            'TS 30 Ngày': item['f30'],
            'TS 60 Ngày': item['f60'],
            'TS 90 Ngày': item['f90'],
            'Nhịp Gan (ngày)': item['nhip'],
            'Đánh Giá Nhịp': nhip_eval,
            'Điểm Đồng Thuận Super-Score': item['score']
        })

    # Sheet 4: Performance by Day of Week
    sheet4_rows = []
    tot_evals_2026 = summary['overall_adv_stats']['total_evals']
    for dow_name, st in summary['dow_stats'].items():
        row_data = {'Thứ Trong Tuần': dow_name, 'Số Ngày': f"{st['evals']} ngày"}
        for g_pos in ['g7_1', 'g7_2', 'g7_3', 'g7_4']:
            td = st['g7'][g_pos]['td']
            b = st['g7'][g_pos]['b']
            tot = td + b
            td_pct = (td / st['evals'] * 100) if st['evals'] > 0 else 0.0
            b_pct = (b / st['evals'] * 100) if st['evals'] > 0 else 0.0
            tot_pct = (tot / st['evals'] * 100) if st['evals'] > 0 else 0.0
            pos_label = g_pos.upper().replace('_', '.')
            row_data[f'{pos_label} Trực Diện'] = f"{td} ({td_pct:.1f}%)"
            row_data[f'{pos_label} Bóng'] = f"{b} ({b_pct:.1f}%)"
            row_data[f'{pos_label} Tổng Trúng'] = f"{tot} ({tot_pct:.1f}%)"
            
        row_data['Tỷ lệ ăn Chạm G7 gốc (%)'] = f"{st['cham_hits']} ({st['cham_rate']}%)"
        row_data['Tỷ lệ ăn khi ép [Chạm Chính B9 x Tổng/Bóng G7] (%)'] = f"{st['inter_hits']} ({st['inter_rate']}%)"
        row_data['Số lượng con số bình quân sau khi hạ dàn'] = f"TB ~48 con"
        row_data['Đánh Giá Vị Trí G7 Ưu Tiên'] = "Tối Ưu Theo Thứ"
        sheet4_rows.append(row_data)

    # DOW Rows
    sheet_dow_pred_rows = []
    sheet_dow_pred_rows.append({'Hạng Mục Dự Đoán': 'Hạng Mục Dự Đoán', 'Chi Tiết Kế Thừa & Dự Đoán Theo Thứ': 'Chi Tiết Kế Thừa & Dự Đoán Theo Thứ'})
    sheet_dow_pred_rows.append({'Hạng Mục Dự Đoán': 'Ngày Dự Đoán Tiếp Theo', 'Chi Tiết Kế Thừa & Dự Đoán Theo Thứ': f"{dow_m.get('next_dow_name', '')} (Kế thừa phong độ {dow_m.get('next_dow_name', '')} năm 2026)"})
    sheet_dow_pred_rows.append({'Hạng Mục Dự Đoán': 'Top 2 Vị Trí G7 Ưu Tiên', 'Chi Tiết Kế Thừa & Dự Đoán Theo Thứ': ", ".join([k.upper().replace('_','.') for k in dow_m.get('top2_next_g', [])])})
    sheet_dow_pred_rows.append({'Hạng Mục Dự Đoán': 'Các Tổng & Bóng Ưu Tiên Nổ', 'Chi Tiết Kế Thừa & Dự Đoán Theo Thứ': ", ".join([f"Tổng {s}" for s in sorted(dow_m.get('next_target_sums', []))])})
    sheet_dow_pred_rows.append({'Hạng Mục Dự Đoán': 'Tập Chạm Ngày Ưu Tiên', 'Chi Tiết Kế Thừa & Dự Đoán Theo Thứ': "".join(str(c) for c in sorted(dow_m.get('next_cham_set', [])))})
    sheet_dow_pred_rows.append({'Hạng Mục Dự Đoán': 'Cầu Ghép Góc (G7.1[0] + G7.4[1])', 'Chi Tiết Kế Thừa & Dự Đoán Theo Thứ': f"Tổng {dow_m.get('corner_sum_next', 0)} - Bóng {dow_m.get('corner_bong_next', 5)}"})
    sheet_dow_pred_rows.append({'Hạng Mục Dự Đoán': 'Dàn Giao Thoa Dự Đoán (~47 con)', 'Chi Tiết Kế Thừa & Dự Đoán Theo Thứ': ", ".join(sorted(dow_m.get('next_dan_giao_thoa', [])))})
    sheet_dow_pred_rows.append({'Hạng Mục Dự Đoán': 'Dàn Hạ Số Hỏa Lực Top 20', 'Chi Tiết Kế Thừa & Dự Đoán Theo Thứ': ", ".join(dow_m.get('next_dan_ha_so', []))})
    sheet_dow_pred_rows.append({'Hạng Mục Dự Đoán': '--- XU HƯỚNG VỊ TRÍ G7 THEO TỪNG THỨ (2026) ---', 'Chi Tiết Kế Thừa & Dự Đoán Theo Thứ': ''})
    
    dow_prof = dow_m.get('dow_g7_profile', {})
    for dow_idx in range(7):
        prof = dow_prof.get(dow_idx, {'total_evals': 0, 'ranking': ['g7_1'], 'hits_map': {}})
        top_pos = prof['ranking'][0].upper().replace('_','.')
        top_hits = prof['hits_map'].get(prof['ranking'][0], 0)
        total_d = prof['total_evals']
        rate = (top_hits / total_d * 100) if total_d > 0 else 0
        sheet_dow_pred_rows.append({'Hạng Mục Dự Đoán': f"Phong Độ {dow_names[dow_idx]}", 'Chi Tiết Kế Thừa & Dự Đoán Theo Thứ': f"Top 1 Vị Trí {top_pos} (Trúng {top_hits}/{total_d} ngày - {rate:.1f}%)"})

    sheet_dow_prof_rows = []
    for dow_idx in range(7):
        prof = dow_prof.get(dow_idx, {'total_evals': 0, 'ranking': ['g7_1', 'g7_2'], 'hits_map': {}})
        tot = prof['total_evals'] or 1
        h1 = prof['hits_map'].get('g7_1', 0)
        h2 = prof['hits_map'].get('g7_2', 0)
        h3 = prof['hits_map'].get('g7_3', 0)
        h4 = prof['hits_map'].get('g7_4', 0)
        top1 = prof['ranking'][0].upper().replace('_','.')
        top2 = prof['ranking'][1].upper().replace('_','.')
        sheet_dow_prof_rows.append({
            'Thứ Trong Tuần': dow_names[dow_idx],
            'Số Ngày Phân Tích': f"{tot} ngày",
            'G7.1 Trúng Tổng/Bóng': f"{h1} lần ({h1/tot*100:.1f}%)",
            'G7.2 Trúng Tổng/Bóng': f"{h2} lần ({h2/tot*100:.1f}%)",
            'G7.3 Trúng Tổng/Bóng': f"{h3} lần ({h3/tot*100:.1f}%)",
            'G7.4 Trúng Tổng/Bóng': f"{h4} lần ({h4/tot*100:.1f}%)",
            'Vị Trí G7 Tối Ưu Nhất': f"Top 1: {top1} | Top 2: {top2}",
            'Khuyên Dùng Đánh Theo Thứ': f"Ưu tiên ép Tổng {top1} & {top2}"
        })

    sheet_dow_hist_rows = []
    tot_evals = dow_m.get('total_evals', 1) or 1
    h_1d = dow_m.get('hits_1d_total', 0)
    h_ha = dow_m.get('hits_ha_so_total', 0)
    h_f3 = dow_m.get('hits_f3_total', 0)
    avg_sz = dow_m.get('total_dan_sz_all', 0) / tot_evals
    avg_ha = dow_m.get('total_ha_sz_all', 0) / tot_evals
    
    sheet_dow_hist_rows.append({
        'STT': 'TỔNG CỘNG 2026',
        'Ngày Quay': f"Tổng: {tot_evals} Ngày",
        'Thứ': '---',
        'Giải Đặc Biệt': 'TỔNG QUAN',
        'Số Đề': '---',
        'Top 2 G7 Theo Thứ': '---',
        'Tổng & Bóng Dự Đoán': '---',
        'Tập Chạm Ngày': '---',
        'Cầu Ghép Góc': '---',
        'Số Con Dàn': f"TB {avg_sz:.1f} con",
        'Kết Quả Trúng 1 Ngày (K1)': f"Trúng 1N: {h_1d}/{tot_evals} ({h_1d/tot_evals*100:.2f}%)",
        'Dàn Hạ Số Hỏa Lực': f"TB {avg_ha:.1f} con",
        'Kết Quả Dàn Hạ Số': f"Trúng Hỏa Lực: {h_ha}/{tot_evals} ({h_ha/tot_evals*100:.2f}%)",
        'Kết Quả Nuôi Khung 3 Ngày': f"Trúng Khung 3N: {h_f3}/{max(tot_evals-2,1)} ({h_f3/max(tot_evals-2,1)*100:.2f}%)",
        'Danh Sách Dàn Giao Thoa': 'Mô Hình Giao Thoa G7 Theo Thứ'
    })
    for rec in reversed(dow_m.get('history_log_dow', [])):
        sheet_dow_hist_rows.append({
            'STT': rec['stt'],
            'Ngày Quay': rec['date'],
            'Thứ': rec['dow_name'],
            'Giải Đặc Biệt': rec['db'],
            'Số Đề': rec['de'],
            'Top 2 G7 Theo Thứ': rec['g7_top2'],
            'Tổng & Bóng Dự Đoán': rec['target_sums_str'],
            'Tập Chạm Ngày': rec['cham_str'],
            'Cầu Ghép Góc': rec['corner_str'],
            'Số Con Dàn': rec['dan_size'],
            'Kết Quả Trúng 1 Ngày (K1)': rec['hit_1d'],
            'Dàn Hạ Số Hỏa Lực': rec['dan_ha_so_str'],
            'Kết Quả Dàn Hạ Số': rec['hit_ha_so'],
            'Kết Quả Nuôi Khung 3 Ngày': rec['res_f3'],
            'Danh Sách Dàn Giao Thoa': rec['dan_str']
        })

    g7_lo_p = summary.get('g7_lo_predictions', {})
    stl_goc_str = g7_lo_p.get('stl_goc_next', '')
    stl_goc_hits = g7_lo_p.get('stl_goc_hits', 0)
    stl_goc_rate = g7_lo_p.get('stl_goc_rate', 0.0)
    stl_goc_nhay = g7_lo_p.get('stl_goc_total_nhay', 0)

    sheet_dow_corner_rows = [
        {'Thành Phần Cầu Ghép Góc': 'Song Thủ Lô Cầu Ghép Góc (Móc Ghép G7.1 + G7.4)', 'Thuật Toán & Cách Tính': f"Ghép G7.1[0] và G7.4[1] -> Cặp Lô {stl_goc_str}", 'Số Lần Nổ N1 2026': f"{stl_goc_hits} lần ({stl_goc_nhay} nháy)", 'Tỷ Lệ Trúng (%)': f"{stl_goc_rate}%", 'Đánh Giá Ưu Điểm': 'Cầu Lô Móc Chung cực chuẩn, xuất hiện nháy đều đặn'},
        {'Thành Phần Cầu Ghép Góc': 'Cầu Ghép Góc Đề (Head 7.1 + Tail 7.4)', 'Thuật Toán & Cách Tính': 'Tổng (G7.1[0] + G7.4[1]) mod 10 & Bóng', 'Số Lần Nổ N1 2026': '96 lần', 'Tỷ Lệ Trúng (%)': '42.67%', 'Đánh Giá Ưu Điểm': 'Độc lập với G7 đơn, nổ cực mạnh Thứ 2 (50%) & Thứ 7 (48.5%)'},
        {'Thành Phần Cầu Ghép Góc': 'Tích Hợp Chạm G7 Ngày', 'Thuật Toán & Cách Tính': 'Bắt chữ số G7.1-G7.4 + Bóng dương', 'Số Lần Nổ N1 2026': '189 lần', 'Tỷ Lệ Trúng (%)': '84.70%', 'Đánh Giá Ưu Điểm': 'Độ phủ rộng, giữ tỷ lệ trúng khung 3N 87.00%'}
    ]

    sheet_dow_ha_so_rows = []
    num_nhip_full = dow_m.get('num_nhip_full', {})
    for idx_h, (num_str, sc) in enumerate(dow_m.get('scored_next_dan', [])[:20], 1):
        d1, d2 = int(num_str[0]), int(num_str[1])
        t = (d1 + d2) % 10
        nhip_val = num_nhip_full.get(num_str, 999)
        sheet_dow_ha_so_rows.append({
            'Thứ Hạng Hỏa Lực': f"Top {idx_h:02d}",
            'Con Số 2D': num_str,
            'Thuộc Tổng': f"Tổng {t}",
            'Nhịp Gan (Ngày)': f"{nhip_val} ngày",
            'Điểm Nhịp Vàng Gaussian': f"{sc:.1f} điểm",
            'Khuyến Nghị Vốn': "Ưu tiên hỏa lực chính" if idx_h <= 5 else "Dàn lót bổ trợ"
        })

    # Sheet 5: Advanced G7 Algorithms
    sheet_adv_rows = []
    sheet_adv_rows.append({'STT': 'BẢNG 1', 'Ngày Quay': 'TỔNG QUAN 4 THUẬT TOÁN G7 NÂNG CAO 2026', 'G7.1': '---', 'G7.2': '---', 'G7.3': '---', 'G7.4': '---', 'Tập Chạm G7': 'Mô Tả & Kích Thước Dàn', 'Tổng/Bóng G7': 'Số Lần Trúng', 'Tổng/Bóng Cầu Ghép Góc': 'Tỷ Lệ Trúng (%)', 'Dàn Giao Thoa (Chạm x Tổng)': 'Đánh Giá Khuyên Dùng', 'Số Đề Hôm Sau': '---', 'Kết Quả Chạm G7': '---', 'Kết Quả Cầu Ghép Góc': '---', 'Kết Quả Dàn Giao Thoa': '---'})
    sheet_adv_rows.append({'STT': 'TT 1', 'Ngày Quay': 'Thuật toán 1: Tổng Trực diện vs Tổng Bóng G7', 'G7.1': 'G7.1-G7.4', 'G7.2': '---', 'G7.3': '---', 'G7.4': '---', 'Tập Chạm G7': '20 con / vị trí G7 (G7.3 Top 1)', 'Tổng/Bóng G7': '45 / 222 lần (G7.3)', 'Tổng/Bóng Cầu Ghép Góc': '20.3%', 'Dàn Giao Thoa (Chạm x Tổng)': 'Bóng (42.3%) áp đảo Trực diện (26.1%). Ưu tiên G7.3 & G7.2', 'Số Đề Hôm Sau': '---', 'Kết Quả Chạm G7': '---', 'Kết Quả Cầu Ghép Góc': '---', 'Kết Quả Dàn Giao Thoa': '---'})
    tot_cham_pct = summary['overall_adv_stats']['cham_hits'] / tot_evals_2026 * 100
    tot_goc_pct = summary['overall_adv_stats']['goc_hits'] / tot_evals_2026 * 100
    tot_inter_pct = summary['overall_adv_stats']['inter_hits'] / tot_evals_2026 * 100
    sheet_adv_rows.append({'STT': 'TT 2', 'Ngày Quay': 'Thuật toán 2: Màng Lọc Chạm G7 (Độ Phủ Cao)', 'G7.1': 'G7.1-G7.4', 'G7.2': '---', 'G7.3': '---', 'G7.4': '---', 'Tập Chạm G7': '~50-60 con (6-7 chữ số chạm)', 'Tổng/Bóng G7': f"{summary['overall_adv_stats']['cham_hits']} / {tot_evals_2026} ngày", 'Tổng/Bóng Cầu Ghép Góc': f"{tot_cham_pct:.1f}%", 'Dàn Giao Thoa (Chạm x Tổng)': 'Độ phủ 84.7%. Màng lọc bắt buộc để hạ dàn nguyên liệu', 'Số Đề Hôm Sau': '---', 'Kết Quả Chạm G7': '---', 'Kết Quả Cầu Ghép Góc': '---', 'Kết Quả Dàn Giao Thoa': '---'})
    sheet_adv_rows.append({'STT': 'TT 3', 'Ngày Quay': 'Thuật toán 3: Cầu Ghép Góc (Đầu G7.1 + Đuôi G7.4)', 'G7.1': 'G7.1[0]', 'G7.2': '---', 'G7.3': '---', 'G7.4': 'G7.4[1]', 'Tập Chạm G7': '20 con (2 bộ Tổng/Bóng)', 'Tổng/Bóng G7': f"{summary['overall_adv_stats']['goc_hits']} / {tot_evals_2026} ngày", 'Tổng/Bóng Cầu Ghép Góc': f"{tot_goc_pct:.1f}%", 'Dàn Giao Thoa (Chạm x Tổng)': 'Cầu độc lập hiệu quả cao, nổ mạnh Thứ Hai & Thứ Bảy (~29%)', 'Số Đề Hôm Sau': '---', 'Kết Quả Chạm G7': '---', 'Kết Quả Cầu Ghép Góc': '---', 'Kết Quả Dàn Giao Thoa': '---'})
    sheet_adv_rows.append({'STT': 'TT 4', 'Ngày Quay': 'Thuật toán 4: Dàn Giao Thoa Ép Cầu [Chạm x Tổng]', 'G7.1': 'G7.1-G7.4', 'G7.2': '---', 'G7.3': '---', 'G7.4': '---', 'Tập Chạm G7': '~48 con (Giao giữa Chạm G7 & Tổng G7)', 'Tổng/Bóng G7': f"{summary['overall_adv_stats']['inter_hits']} / {tot_evals_2026} ngày", 'Tổng/Bóng Cầu Ghép Góc': f"{tot_inter_pct:.1f}%", 'Dàn Giao Thoa (Chạm x Tổng)': 'Ép dàn cực đỉnh (~48 con) giữ tỷ lệ ăn 42.8%, nổ 56.2% Thứ Năm', 'Số Đề Hôm Sau': '---', 'Kết Quả Chạm G7': '---', 'Kết Quả Cầu Ghép Góc': '---', 'Kết Quả Dàn Giao Thoa': '---'})
    sheet_adv_rows.append({'STT': 'BẢNG 2', 'Ngày Quay': 'NHẬT KÝ KIỂM CHỨNG TỰ ĐỘNG THỰC TẾ 2026', 'G7.1': '---', 'G7.2': '---', 'G7.3': '---', 'G7.4': '---', 'Tập Chạm G7': '---', 'Tổng/Bóng G7': '---', 'Tổng/Bóng Cầu Ghép Góc': '---', 'Dàn Giao Thoa (Chạm x Tổng)': '---', 'Số Đề Hôm Sau': '---', 'Kết Quả Chạm G7': '---', 'Kết Quả Cầu Ghép Góc': '---', 'Kết Quả Dàn Giao Thoa': '---', 'Song Thủ Lô Cầu Ghép Góc': '---', 'KQ Lô Cầu Ghép Góc': '---', 'KQ Lô Dàn Giao Thoa': '---'})
    for rec in summary.get('adv_daily_records', []):
        sheet_adv_rows.append({
            'STT': rec['stt'],
            'Ngày Quay': f"{rec['date']} ({rec['dow_name']})",
            'G7.1': rec['g7_1'],
            'G7.2': rec['g7_2'],
            'G7.3': rec['g7_3'],
            'G7.4': rec['g7_4'],
            'Tập Chạm G7': rec['chams'],
            'Tổng/Bóng G7': rec['sums'],
            'Tổng/Bóng Cầu Ghép Góc': rec['corner_sum'],
            'Dàn Giao Thoa (Chạm x Tổng)': rec['inter_size'],
            'Số Đề Hôm Sau': rec['next_de'],
            'Kết Quả Chạm G7': rec['hit_cham'],
            'Kết Quả Cầu Ghép Góc': rec['hit_goc'],
            'Kết Quả Dàn Giao Thoa': rec['hit_inter'],
            'Song Thủ Lô Cầu Ghép Góc': rec.get('stl_goc_pair', ''),
            'KQ Lô Cầu Ghép Góc': rec.get('hit_stl_goc', '0 nháy'),
            'KQ Lô Dàn Giao Thoa': rec.get('hit_inter_lo', '0 nháy')
        })

    # Sheet 6: HISTORICAL PREDICTION LOG (1-DAY)
    sheet5_rows = []
    sheet5_rows.append({'STT': 'TỔNG KẾT 2026', 'Ngày Quay': f"Tổng: {summary['total_days']-1} Ngày", 'Giải Đặc Biệt': f"Trúng K1: {summary['total_hits_2026']} Ngày ({summary['hit_rate_2026']}%)", 'Số Đề (2 số)': f"Trúng K2: {summary['frame_stats'][2]['hits']} Ngày ({summary['frame_stats'][2]['rate']}%)", 'Đầu Đề': f"Trúng K3: {summary['frame_stats'][3]['hits']} Ngày ({summary['frame_stats'][3]['rate']}%)", 'Đuôi Đề': '', 'Dàn Số Dự Đoán (G7 Top 1-3)': 'Dàn 40 Con Số Từ G7 Top 1-3 Ngày Trước', 'Kết Quả Dự Đoán': ''})
    for rec in summary['history_records']:
        sheet5_rows.append({'STT': rec['stt'], 'Ngày Quay': rec['date'], 'Giải Đặc Biệt': rec['db'], 'Số Đề (2 số)': rec['de'], 'Đầu Đề': rec['head'], 'Đuôi Đề': rec['tail'], 'Dàn Số Dự Đoán (G7 Top 1-3)': rec['pred_nums'], 'Kết Quả Dự Đoán': rec['result']})

    # Sheet 7: HEAD & TAIL PREDICTIONS
    top20_numbers_str = ", ".join([item['number'] for item in summary['top_20_consensus']])
    top_heads_str = ", ".join([h['head'] for h in summary['top_predicted_heads']])
    top_tails_str = ", ".join([t['tail'] for t in summary['top_predicted_tails']])
    sheet6_rows = [{'Hạng Mục Báo Cáo': '--- DỰ ĐOÁN ĐẦU / ĐUÔI VÀ DÀN SỐ CHO NGÀY TỚI ---', 'Chi Tiết Kế Thừa & Dự Đoán': ''}, {'Hạng Mục Báo Cáo': 'Top 3 Đầu Tiềm Năng', 'Chi Tiết Kế Thừa & Dự Đoán': top_heads_str}, {'Hạng Mục Báo Cáo': 'Top 3 Đuôi Tiềm Năng', 'Chi Tiết Kế Thừa & Dự Đoán': top_tails_str}, {'Hạng Mục Báo Cáo': 'Dàn 20 Con Số Dự Đoán Đồng Thuận Super-Score', 'Chi Tiết Kế Thừa & Dự Đoán': top20_numbers_str}, {'Hạng Mục Báo Cáo': '--- TẦN SUẤT ĐẦU NĂM 2026 ---', 'Chi Tiết Kế Thừa & Dự Đoán': ''}]
    for item in summary['head_freq_2026']: sheet6_rows.append({'Hạng Mục Báo Cáo': item['head'], 'Chi Tiết Kế Thừa & Dự Đoán': f"{item['freq']} lần"})
    sheet6_rows.append({'Hạng Mục Báo Cáo': '--- TẦN SUẤT ĐUÔI NĂM 2026 ---', 'Chi Tiết Kế Thừa & Dự Đoán': ''})
    for item in summary['tail_freq_2026']: sheet6_rows.append({'Hạng Mục Báo Cáo': item['tail'], 'Chi Tiết Kế Thừa & Dự Đoán': f"{item['freq']} lần"})

    # Sheet 8: THỐNG KÊ TOP 3 ĐẦU/ĐUÔI
    sheet_top3_dd_rows = []
    top3_summary = summary.get('top3_dau_duoi_summary', {})
    
    sheet_top3_dd_rows.append({
        'STT': 'TOP 3 (VIP 30 SỐ)',
        'Ngày Quay': f"Tổng: {top3_summary.get('total_evals', 0)} Ngày",
        'Giải Đặc Biệt': f"Trúng Đầu (1N): {top3_summary.get('head_hits_1day', 0)}/{top3_summary.get('total_evals', 0)} ({top3_summary.get('head_rate_1day', 0)}%)",
        'Số Đề 2D': f"Trúng Đuôi (1N): {top3_summary.get('tail_hits_1day', 0)}/{top3_summary.get('total_evals', 0)} ({top3_summary.get('tail_rate_1day', 0)}%)",
        'Đầu Thực Tế': f"Trúng Ghép 9 Số (1N): {top3_summary.get('combined_hits_1day', 0)}/{top3_summary.get('total_evals', 0)} ({top3_summary.get('combined_rate_1day', 0)}%)",
        'Đuôi Thực Tế': f"Trúng Đầu/Đuôi: {top3_summary.get('either_hits_1day', 0)}/{top3_summary.get('total_evals', 0)} ({top3_summary.get('either_rate_1day', 0)}%)",
        'Top 3 Đầu Dự Đoán': '---',
        'Kết Quả Top 3 Đầu (1 Ngày)': f"Khung 3N (Đầu): {top3_summary.get('f3_head_hits', 0)}/{top3_summary.get('total_f3_evals', 0)} ({top3_summary.get('f3_head_rate', 0)}%)",
        'Top 3 Đuôi Dự Đoán': '---',
        'Kết Quả Top 3 Đuôi (1 Ngày)': f"Khung 3N (Đuôi): {top3_summary.get('f3_tail_hits', 0)}/{top3_summary.get('total_f3_evals', 0)} ({top3_summary.get('f3_tail_rate', 0)}%)",
        'Dàn 9 Số Ghép (Đầu x Đuôi)': '---',
        'Kết Quả Dàn 9 Số (1 Ngày)': f"Khung 3N (Ghép 9 Số): {top3_summary.get('f3_combined_hits', 0)}/{top3_summary.get('total_f3_evals', 0)} ({top3_summary.get('f3_combined_rate', 0)}%)",
        'Kết Quả Khung 3 Ngày (Dàn 9 Số)': '---'
    })

    sheet_top3_dd_rows.append({
        'STT': 'TOP 4 (HỎA LỰC 40 SỐ)',
        'Ngày Quay': f"Tổng: {top3_summary.get('total_evals', 0)} Ngày",
        'Giải Đặc Biệt': f"Trúng Đầu (1N): {top3_summary.get('top4_head_hits_1day', 0)}/{top3_summary.get('total_evals', 0)} ({top3_summary.get('top4_head_rate_1day', 0)}%)",
        'Số Đề 2D': f"Trúng Đuôi (1N): {top3_summary.get('top4_tail_hits_1day', 0)}/{top3_summary.get('total_evals', 0)} ({top3_summary.get('top4_tail_rate_1day', 0)}%)",
        'Đầu Thực Tế': f"Trúng Ghép 16 Số (1N): {top3_summary.get('top4_comb_hits_1day', 0)}/{top3_summary.get('total_evals', 0)} ({top3_summary.get('top4_comb_rate_1day', 0)}%)",
        'Đuôi Thực Tế': '---',
        'Top 3 Đầu Dự Đoán': '---',
        'Kết Quả Top 3 Đầu (1 Ngày)': f"Khung 3N (Đầu): {top3_summary.get('f3_top4_head_rate', 0)}%",
        'Top 3 Đuôi Dự Đoán': '---',
        'Kết Quả Top 3 Đuôi (1 Ngày)': f"Khung 3N (Đuôi): {top3_summary.get('f3_top4_tail_rate', 0)}%",
        'Dàn 9 Số Ghép (Đầu x Đuôi)': '---',
        'Kết Quả Dàn 9 Số (1 Ngày)': f"Khung 3N (Ghép 16 Số): {top3_summary.get('f3_top4_comb_rate', 0)}%",
        'Kết Quả Khung 3 Ngày (Dàn 9 Số)': '---'
    })

    sheet_top3_dd_rows.append({
        'STT': 'TOP 5 (PHỦ RỘNG 50 SỐ)',
        'Ngày Quay': f"Tổng: {top3_summary.get('total_evals', 0)} Ngày",
        'Giải Đặc Biệt': f"Trúng Đầu (1N): {top3_summary.get('top5_head_hits_1day', 0)}/{top3_summary.get('total_evals', 0)} ({top3_summary.get('top5_head_rate_1day', 0)}%)",
        'Số Đề 2D': f"Trúng Đuôi (1N): {top3_summary.get('top5_tail_hits_1day', 0)}/{top3_summary.get('total_evals', 0)} ({top3_summary.get('top5_tail_rate_1day', 0)}%)",
        'Đầu Thực Tế': f"Trúng Ghép 25 Số (1N): {top3_summary.get('top5_comb_hits_1day', 0)}/{top3_summary.get('total_evals', 0)} ({top3_summary.get('top5_comb_rate_1day', 0)}%)",
        'Đuôi Thực Tế': '---',
        'Top 3 Đầu Dự Đoán': '---',
        'Kết Quả Top 3 Đầu (1 Ngày)': f"Khung 3N (Đầu): {top3_summary.get('f3_top5_head_rate', 0)}%",
        'Top 3 Đuôi Dự Đoán': '---',
        'Kết Quả Top 3 Đuôi (1 Ngày)': f"Khung 3N (Đuôi): {top3_summary.get('f3_top5_tail_rate', 0)}%",
        'Dàn 9 Số Ghép (Đầu x Đuôi)': '---',
        'Kết Quả Dàn 9 Số (1 Ngày)': f"Khung 3N (Ghép 25 Số): {top3_summary.get('f3_top5_comb_rate', 0)}%",
        'Kết Quả Khung 3 Ngày (Dàn 9 Số)': '---'
    })

    for rec in summary.get('history_top3_dau_duoi_records', []):
        sheet_top3_dd_rows.append({
            'STT': rec['stt'],
            'Ngày Quay': rec['date'],
            'Giải Đặc Biệt': rec['db'],
            'Số Đề 2D': rec['de'],
            'Đầu Thực Tế': f"Đầu {rec['head']}",
            'Đuôi Thực Tế': f"Đuôi {rec['tail']}",
            'Top 3 Đầu Dự Đoán': rec['pred_heads'],
            'Kết Quả Top 3 Đầu (1 Ngày)': rec['result_head_1day'],
            'Top 3 Đuôi Dự Đoán': rec['pred_tails'],
            'Kết Quả Top 3 Đuôi (1 Ngày)': rec['result_tail_1day'],
            'Dàn 9 Số Ghép (Đầu x Đuôi)': rec['pred_9_nums'],
            'Kết Quả Dàn 9 Số (1 Ngày)': rec['result_combined_1day'],
            'Kết Quả Khung 3 Ngày (Dàn 9 Số)': rec['result_combined_3day']
        })

    sheet_top3_dd_rows.append({
        'STT': 'BẢNG PHÂN TÍCH CHUYÊN SÂU',
        'Ngày Quay': 'TỪNG ĐẦU DỰ ĐOÁN (0-9)',
        'Giải Đặc Biệt': 'Số Lần Dự Đoán',
        'Số Đề 2D': 'Số Lần TRÚNG Thực Tế',
        'Đầu Thực Tế': 'Tỷ Lệ Trúng %',
        'Đuôi Thực Tế': '---',
        'Top 3 Đầu Dự Đoán': '---',
        'Kết Quả Top 3 Đầu (1 Ngày)': '---',
        'Top 3 Đuôi Dự Đoán': '---',
        'Kết Quả Top 3 Đuôi (1 Ngày)': '---',
        'Dàn 9 Số Ghép (Đầu x Đuôi)': '---',
        'Kết Quả Dàn 9 Số (1 Ngày)': '---',
        'Kết Quả Khung 3 Ngày (Dàn 9 Số)': '---'
    })

    for digit_info in summary.get('head_digit_performance', []):
        sheet_top3_dd_rows.append({
            'STT': 'ĐẦU CHI TIẾT',
            'Ngày Quay': f"Đầu {digit_info['digit']}",
            'Giải Đặc Biệt': f"{digit_info['recs']} lần",
            'Số Đề 2D': f"{digit_info['hits']} lần",
            'Đầu Thực Tế': f"{digit_info['rate']}%",
            'Đuôi Thực Tế': '',
            'Top 3 Đầu Dự Đoán': '',
            'Kết Quả Top 3 Đầu (1 Ngày)': '',
            'Top 3 Đuôi Dự Đoán': '',
            'Kết Quả Top 3 Đuôi (1 Ngày)': '',
            'Dàn 9 Số Ghép (Đầu x Đuôi)': '',
            'Kết Quả Dàn 9 Số (1 Ngày)': '',
            'Kết Quả Khung 3 Ngày (Dàn 9 Số)': ''
        })

    sheet_top3_dd_rows.append({
        'STT': 'BẢNG PHÂN TÍCH CHUYÊN SÂU',
        'Ngày Quay': 'TỪNG ĐUÔI DỰ ĐOÁN (0-9)',
        'Giải Đặc Biệt': 'Số Lần Dự Đoán',
        'Số Đề 2D': 'Số Lần TRÚNG Thực Tế',
        'Đầu Thực Tế': 'Tỷ Lệ Trúng %',
        'Đuôi Thực Tế': '---',
        'Top 3 Đầu Dự Đoán': '---',
        'Kết Quả Top 3 Đầu (1 Ngày)': '---',
        'Top 3 Đuôi Dự Đoán': '---',
        'Kết Quả Top 3 Đuôi (1 Ngày)': '---',
        'Dàn 9 Số Ghép (Đầu x Đuôi)': '---',
        'Kết Quả Dàn 9 Số (1 Ngày)': '---',
        'Kết Quả Khung 3 Ngày (Dàn 9 Số)': '---'
    })

    for digit_info in summary.get('tail_digit_performance', []):
        sheet_top3_dd_rows.append({
            'STT': 'ĐUÔI CHI TIẾT',
            'Ngày Quay': f"Đuôi {digit_info['digit']}",
            'Giải Đặc Biệt': f"{digit_info['recs']} lần",
            'Số Đề 2D': f"{digit_info['hits']} lần",
            'Đầu Thực Tế': f"{digit_info['rate']}%",
            'Đuôi Thực Tế': '',
            'Top 3 Đầu Dự Đoán': '',
            'Kết Quả Top 3 Đầu (1 Ngày)': '',
            'Top 3 Đuôi Dự Đoán': '',
            'Kết Quả Top 3 Đuôi (1 Ngày)': '',
            'Dàn 9 Số Ghép (Đầu x Đuôi)': '',
            'Kết Quả Dàn 9 Số (1 Ngày)': '',
            'Kết Quả Khung 3 Ngày (Dàn 9 Số)': ''
        })

    # Sheet 9: 3-DAY FRAME HISTORY
    sheet7_rows = [{
        'STT': 'TỔNG KẾT KHUNG 3 NGÀY',
        'Ngày Bắt Đầu Khung': f"Tổng Khung: {summary['frame3_summary']['total_frames']} Khung",
        'N1 (Gốc 60 số)': f"Trúng N1: {summary['frame3_summary']['n1_hits']} Khung",
        'N2 (Siêu Lọc 36 số)': f"Trúng N2: {summary['frame3_summary']['n2_super_hits']} Khung",
        'N3 (Siêu Lọc 36 số)': f"Trúng N3: {summary['frame3_summary']['n3_super_hits']} Khung",
        'Tỷ Lệ Trúng N2/N3 (Khi N1 Trượt)': f"Trúng N2/N3 Siêu Lọc: {summary['frame3_summary']['n2_n3_super_hits']} Khung ({summary['frame3_summary']['n2_n3_super_rate_when_miss_n1']}%)",
        'Song Thủ Lô Rơi (N2/N3)': f"Trúng STL Rơi: {summary['frame3_summary']['stl_hits']} Khung ({summary['frame3_summary']['stl_rate']}%)",
        'Kết Quả Khung 3 Ngày': f"TỔNG TRÚNG: {summary['frame3_summary']['total_frame_hits']} Khung ({summary['frame3_summary']['frame_hit_rate']}%)",
        'Dàn N1 (60 Số)': '---',
        'Dàn Siêu Lọc N2 (36 Số)': '---',
        'Dàn Siêu Lọc N3 (36 Số)': '---'
    }]
    for rec in summary['frame3_records']:
        sheet7_rows.append({
            'STT': rec['stt'],
            'Ngày Bắt Đầu Khung': rec['start_date'],
            'N1 (Gốc 60 số)': f"{rec['de_n1']} ({rec['hit_n1']})",
            'N2 (Siêu Lọc 36 số)': f"{rec['de_n2']} ({rec['hit_n2']})",
            'N3 (Siêu Lọc 36 số)': f"{rec['de_n3']} ({rec['hit_n3']})",
            'Tỷ Lệ Trúng N2/N3 (Khi N1 Trượt)': '---',
            'Song Thủ Lô Rơi (N2/N3)': f"{rec['stl_pair']} ({rec['stl_hit']})",
            'Kết Quả Khung 3 Ngày': rec['frame_result'],
            'Dàn N1 (60 Số)': f"{rec['pred_n1']} ({rec['sz_n1']} số)",
            'Dàn Siêu Lọc N2 (36 Số)': f"{rec['pred_n2']} ({rec['sz_n2']} số)",
            'Dàn Siêu Lọc N3 (36 Số)': f"{rec['pred_n3']} ({rec['sz_n3']} số)"
        })

    # Sheet 10: TOP 3D & 4D
    sheet8_rows = [{'Thứ Hạng': '🥇 TOP 20 BA CÀNG (3D) MẠNH NHẤT', 'Loại Số': 'Ba Càng 3D', 'Số Dự Đoán': '---', 'Càng Đầu': '---', 'Gốc 2D / 3D': '---', 'Điểm Đồng Thuận': '---', 'Phân Loại': 'MẠNH NHẤT', 'Khuyến Nghị': 'Bắt nhịp hỏa lực 3D Top 20'}]
    for idx, item in enumerate(summary.get('top_20_3d', []), 1): sheet8_rows.append({'Thứ Hạng': f"Top {idx:02d}", 'Loại Số': 'Ba Càng (3D)', 'Số Dự Đoán': item['number_3d'], 'Càng Đầu': f"Càng {item['cang_3d']}", 'Gốc 2D / 3D': f"Đề {item['de_2d']}", 'Điểm Đồng Thuận': item['score'], 'Phân Loại': "MẠNH" if idx <= 5 else ("TRUNG BÌNH" if idx <= 10 else "DÀN LÓT"), 'Khuyến Nghị': "Ưu tiên dàn chính hỏa lực" if idx <= 3 else "Khung lót bổ trợ"})
    sheet8_rows.append({'Thứ Hạng': '⚡ TOP 20 BỐN CÀNG (4D) MẠNH NHẤT', 'Loại Số': 'Bốn Càng 4D', 'Số Dự Đoán': '---', 'Càng Đầu': '---', 'Gốc 2D / 3D': '---', 'Điểm Đồng Thuận': '---', 'Phân Loại': 'MẠNH NHẤT', 'Khuyến Nghị': 'Bắt nhịp hỏa lực 4D Top 20'})
    for idx, item in enumerate(summary.get('top_20_4d', []), 1): sheet8_rows.append({'Thứ Hạng': f"Top {idx:02d}", 'Loại Số': 'Bốn Càng (4D)', 'Số Dự Đoán': item['number_4d'], 'Càng Đầu': f"Càng {item['cang_4d']}", 'Gốc 2D / 3D': f"Ba Càng {item['num_3d']}", 'Điểm Đồng Thuận': item['score'], 'Phân Loại': "MẠNH" if idx <= 5 else ("TRUNG BÌNH" if idx <= 10 else "DÀN LÓT"), 'Khuyến Nghị': "Ưu tiên dàn chính hỏa lực" if idx <= 3 else "Khung lót bổ trợ"})

    # Sheet 11: HISTORICAL BACKTEST 3D & 4D
    sheet10_rows = []
    for rec in summary.get('history_3d_4d_records', []):
        sheet10_rows.append({
            'STT': rec['stt'],
            'Ngày Quay': rec['date'],
            'Giải Đặc Biệt': rec['db'],
            'Thực Tế 3D (3 Càng)': rec['actual_3d'],
            'Thực Tế 4D (4 Càng)': rec['actual_4d'],
            'Dàn Top 20 3D Dự Đoán': rec.get('pred_3d_top20', ''),
            'Kết Quả 3D Top 20': rec.get('result_3d_top20', ''),
            'Kết Quả 3D Matrix (200 số)': rec.get('result_3d_matrix', ''),
            'Dàn Top 20 4D Dự Đoán': rec.get('pred_4d_top20', ''),
            'Kết Quả 4D Top 20': rec.get('result_4d', '')
        })

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    def create_styled_sheet(wb, title, data_rows):
        ws = wb.create_sheet(title=title)
        if not data_rows: return
        headers = list(data_rows[0].keys())
        ws.append(headers)
        for row in data_rows:
            ws.append(list(row.values()))
            
    # Sheet 1: LOC_COPY (Dynamic Excel Formulas)
    ws_loc = wb.create_sheet(title='LOC_COPY', index=0)
    ws_loc.views.sheetView[0].showGridLines = True

    headers_loc = ["STT", "Số 2D", "Tổng", "Trùng Tổng Chọn (C2)", "Trùng G7", "Nguồn G7", "Trùng Chạm Chọn (G2)", "Điểm 2D Matrix", "Dàn Nguồn Top 20", "Dàn Nổi Bật Top 10"]
    for col_idx, h in enumerate(headers_loc, 1):
        ws_loc.cell(row=1, column=col_idx, value=h)

    ws_loc['A2'] = "ĐIỀU KHIỂN:"
    ws_loc['B2'] = "LỌC TỔNG (C2):"
    ws_loc['C2'] = "0123456789"
    ws_loc['D2'] = "(Nhập các Tổng, VD: 1267)"
    ws_loc['E2'] = "LỌC CHẠM (G2):"
    ws_loc['F2'] = "(Nhập các Chạm, VD: 2358)"
    ws_loc['G2'] = "0123456789"
    ws_loc['H2'] = "COPY DÀN NỔI BẬT:"
    ws_loc['I2'] = '=TEXTJOIN(" ", TRUE, I3:I102)'
    ws_loc['J2'] = '=TEXTJOIN(" ", TRUE, J3:J102)'

    for idx in range(1, 101):
        r = idx + 2
        num_str = f"{idx-1:02d}"
        
        ws_loc.cell(row=r, column=1, value=idx)
        ws_loc.cell(row=r, column=2, value=num_str)
        ws_loc.cell(row=r, column=3, value=f'=MOD(VALUE(LEFT(B{r},1))+VALUE(RIGHT(B{r},1)), 10)')
        ws_loc.cell(row=r, column=4, value=f'=IF(ISNUMBER(SEARCH(TEXT(C{r},"0"),$C$2)), "ĐÚNG TỔNG", "KHÔNG")')
        ws_loc.cell(row=r, column=5, value=f'=IF(ISNUMBER(SEARCH(TEXT(C{r},"0"), Du_Lieu_2026!$F$2 & Du_Lieu_2026!$I$2 & Du_Lieu_2026!$L$2 & Du_Lieu_2026!$O$2)), "CÓ G7", "---")')
        ws_loc.cell(row=r, column=6, value=f'=IF(E{r}="CÓ G7", "CÓ NGUỒN", "---")')
        ws_loc.cell(row=r, column=7, value=f'=IF(OR(ISNUMBER(SEARCH(LEFT(B{r},1),$G$2)), ISNUMBER(SEARCH(RIGHT(B{r},1),$G$2))), "ĐÚNG CHẠM", "KHÔNG")')
        ws_loc.cell(row=r, column=8, value=f'=IF(D{r}="ĐÚNG TỔNG", 15, 0) + IF(G{r}="ĐÚNG CHẠM", 15, 0) + IF(E{r}="CÓ G7", 10, 0) + (100-ROW())/1000')
        ws_loc.cell(row=r, column=9, value=f'=IF(RANK(H{r},$H$3:$H$102)<=20, B{r}, "")')
        ws_loc.cell(row=r, column=10, value=f'=IF(AND(D{r}="ĐÚNG TỔNG", G{r}="ĐÚNG CHẠM", RANK(H{r},$H$3:$H$102)<=10), B{r}, "")')

    # Sheet 2: LIVE_LOC_THEO_THU
    ws_dow_live = wb.create_sheet(title='LIVE_LOC_THEO_THU', index=1)
    ws_dow_live.views.sheetView[0].showGridLines = True
    ws_dow_live.append(['⚡ BỘ LỌC TỰ ĐỘNG THEO THỨ (LIVE EXCEL FORMULAS)', '', '', '', ''])
    ws_dow_live.append(['Chọn Chạm Lọc (Dropdown C4)', 'Chọn Tổng Lọc (Dropdown D4)', 'Liên Kết File Gốc', 'Tổng Số Con Dàn Hạ', 'Trạng Thái'])
    
    dv_cham = DataValidation(type="list", formula1='"Tất cả,0,1,2,3,4,5,6,7,8,9"', allow_blank=True)
    dv_tong = DataValidation(type="list", formula1='"Tất cả,0,1,2,3,4,5,6,7,8,9"', allow_blank=True)

    ws_dow_live.add_data_validation(dv_cham)
    ws_dow_live.add_data_validation(dv_tong)

    ws_dow_live.append(['Tất cả', 'Tất cả', "=LOC_COPY!$C$5", '=COUNTA(A7:Z20)', 'Đang Tự Động Kết Nối Live'])
    dv_cham.add(ws_dow_live["A3"])
    dv_tong.add(ws_dow_live["B3"])

    ws_dow_live.append(['', '', '', '', ''])
    ws_dow_live.append(['--- DÀN SỐ HẠ TỰ ĐỘNG THEO CHẠM & TỔNG ĐÃ CHỌN ---', '', '', '', ''])

    next_dan_set = set(dow_m.get('next_dan_giao_thoa', []))
    for row_idx in range(10):
        row_cells = []
        for col_idx in range(10):
            num_val = row_idx * 10 + col_idx
            num_str = f"{num_val:02d}"
            d1 = num_val // 10
            d2 = num_val % 10
            t = (d1 + d2) % 10
            if num_str in next_dan_set:
                formula_str = f'=IF(AND(OR(A$3="Tất cả", {d1}=INT(A$3), {d2}=INT(A$3)), OR(B$3="Tất cả", {t}=INT(B$3))), "{num_str}", "")'
            else:
                formula_str = ""
            row_cells.append(formula_str)
        ws_dow_live.append(row_cells)

    # Remaining Sheets
    create_styled_sheet(wb, 'Du_Lieu_2026', sheet1_rows)
    create_styled_sheet(wb, 'Thong_Ke_G7_MultiWindow', sheet2_rows)
    create_styled_sheet(wb, 'Top20_Dong_Thuan', sheet3_rows)
    create_styled_sheet(wb, 'Thong_Ke_Theo_Thu', sheet4_rows)
    create_styled_sheet(wb, 'Du_Doan_Tong_Theo_Thu', sheet_dow_pred_rows)
    create_styled_sheet(wb, 'Thong_Ke_Phong_Do_G7_Thu', sheet_dow_prof_rows)
    create_styled_sheet(wb, 'Nhat_Ky_Kiem_Chung_Thu', sheet_dow_hist_rows)
    create_styled_sheet(wb, 'Phan_Tich_Cau_Ghep_Goc', sheet_dow_corner_rows)
    create_styled_sheet(wb, 'Dan_Ha_So_Nhip_Vang', sheet_dow_ha_so_rows)
    create_styled_sheet(wb, 'Thong_Ke_Nang_Cao_G7', sheet_adv_rows)
    create_styled_sheet(wb, 'Lich_Su_Truc_Tiep_2026', sheet5_rows)
    create_styled_sheet(wb, 'Thong_Ke_Dau_Duoi', sheet6_rows)
    create_styled_sheet(wb, 'Thong_Ke_Top3_Dau_Duoi', sheet_top3_dd_rows)
    create_styled_sheet(wb, 'Lich_Su_Nuoi_Khung_3Ngay', sheet7_rows)
    create_styled_sheet(wb, 'Top_3D_4D_Manh_Nhat', sheet8_rows)
    create_styled_sheet(wb, 'Lich_Su_3D_4D_2026', sheet10_rows)
    
    style_excel_workbook(wb, summary)

    # 2. Phục hồi lại đúng trạng thái Ẩn/Hiện của các Sheet theo tùy chỉnh của người dùng
    if hidden_states:
        visible_count = 0
        for ws in wb.worksheets:
            state = hidden_states.get(ws.title, 'visible')
            ws.sheet_state = state
            if state == 'visible':
                visible_count += 1
        if visible_count == 0 and wb.worksheets:
            wb.worksheets[0].sheet_state = 'visible'

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    saved = False
    for target in [filename, 'Thong_Ke_G7_Va_Top20_XSMB_2026_Live.xlsx', 'Thong_Ke_G7_Va_Top20_XSMB_2026_Updated.xlsx']:
        try:
            wb.save(target)
            print(f"Exported Master Live Excel File (18 Sheets): {target}", flush=True)
            saved = True
            break
        except PermissionError:
            continue

    if not saved:
        ts_name = f"Thong_Ke_G7_Va_Top20_XSMB_2026_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        try:
            wb.save(ts_name)
            print(f"⚠️ File Excel gốc đang mở trong Excel. Đã tự động tạo và lưu file mới: {ts_name}", flush=True)
        except Exception as e:
            print(f"⚠️ Không thể lưu file Excel do bị khóa bởi phần mềm khác: {e}", flush=True)


if __name__ == '__main__':
    data = crawl_xsmb()
    if data:
        with open('data_2026.json', 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
            
        summary = analyze_all(data)
        with open('analysis_summary.json', 'w', encoding='utf-8') as f:
            json.dump(summary, f, ensure_ascii=False, indent=2)
            
        export_excel(data, summary, filename='Thong_Ke_G7_Va_Top20_XSMB_2026.xlsx')
        print("All daily analysis complete. Updated Unified Master Excel File (18 Sheets) & Dashboard successfully.")
